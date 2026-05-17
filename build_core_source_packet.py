#!/usr/bin/env python3
"""Build a QC-gated Core source packet from raw_text/*.txt."""

from __future__ import annotations

import argparse
import csv
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

from select_core_excerpts import select_excerpts, setup_logging, word_count


SOURCE_ID_RE = re.compile(r"^(S\d{3})\.txt$", re.IGNORECASE)
EXTRACTION_ERROR_RE = re.compile(
    r"(#NAME\?|#VALUE!|#REF!|\bNULL\b|\bnan\b|Traceback \(most recent call last\)|"
    r"parser traceback|lllIlIl|IlIl|0O0O|�|\x00)",
    re.IGNORECASE,
)
NUMBER_RE = re.compile(r"\b\d+(?:[.,]\d+)?%?\b")
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")

BOILERPLATE_TERMS = (
    "cookie",
    "skip to",
    "search legislation",
    "privacy",
    "copyright",
    "all rights reserved",
    "creative commons",
    "isbn",
    "please cite",
    "photo credit",
    "terms of use",
    "accessibility",
    "contact us",
    "download",
    "share this page",
    "table of contents",
    "contents overview",
    "skip to main content",
    "cookies on",
    "previous next",
    "more resources",
    "access essential accompanying documents",
    "forward-looking statements",
    "risks and uncertainties",
    "actual results",
    "differ materially",
    "cautionary statements",
    "unresolved staff comments",
    "condensed consolidated statements",
    "risk factors",
    "safe harbor",
    "rights and permissions",
    "electronically signed",
    "official journal",
)

SUMMARY_TERMS = (
    "summary",
    "overview",
    "finding",
    "findings",
    "key messages",
    "highlights",
    "results",
    "executive",
    "in brief",
)

CONCLUSION_TERMS = (
    "conclusion",
    "therefore",
    "thus",
    "implies",
    "implication",
    "recommend",
    "recommendation",
    "should",
    "need to",
    "suggests",
    "indicates",
)

COMPARISON_TERMS = (
    "compared",
    "comparison",
    "higher than",
    "lower than",
    "whereas",
    "while",
    "across",
    "between",
    "among",
    "countries",
    "firms",
    "companies",
    "regions",
    "years",
    "rank",
    "ranking",
    "trend",
    "versus",
    "relative to",
    "by country",
    "by sector",
    "by region",
)

COUNTRY_TERMS = (
    "country",
    "countries",
    "australia",
    "austria",
    "belgium",
    "canada",
    "china",
    "denmark",
    "estonia",
    "europe",
    "european union",
    "finland",
    "france",
    "germany",
    "greece",
    "iceland",
    "ireland",
    "italy",
    "japan",
    "latvia",
    "lithuania",
    "mexico",
    "netherlands",
    "norway",
    "poland",
    "portugal",
    "spain",
    "sweden",
    "switzerland",
    "united kingdom",
    "united states",
)


DOMAIN_KEYWORDS = {
    "AI and digital governance": (
        "artificial intelligence",
        " ai ",
        "algorithm",
        "model",
        "digital",
        "open data",
        "data governance",
        "cyber",
        "technology",
    ),
    "Public governance and regulation": (
        "government",
        "public sector",
        "policy",
        "regulation",
        "act 2024",
        "trust",
        "institution",
        "governance",
    ),
    "Economic outlook and development": (
        "gdp",
        "economic",
        "growth",
        "inflation",
        "poverty",
        "development",
        "income",
        "recovery",
    ),
    "Corporate performance and finance": (
        "annual report",
        "financial statements",
        "revenue",
        "operating income",
        "cash flow",
        "shareholders",
        "company",
    ),
    "Health and social outcomes": (
        "health",
        "obesity",
        "disease",
        "mortality",
        "patients",
        "social",
        "well-being",
    ),
    "Energy, climate, and environment": (
        "energy",
        "climate",
        "emissions",
        "greenhouse",
        "environment",
        "renewable",
        "efficiency",
    ),
    "Demographics and population": (
        "population",
        "fertility",
        "mortality",
        "migration",
        "demographic",
        "ageing",
    ),
    "Labor, education, and skills": (
        "labour",
        "labor",
        "employment",
        "education",
        "skills",
        "workers",
        "students",
    ),
}


@dataclass
class SourceMeta:
    source_id: str
    files: list[dict[str, str]]
    canonical_source_id: str
    variant_source_ids: list[str]
    variant_files: list[str]
    duplicate_reason: str


@dataclass
class Flags:
    has_summary: bool
    has_evidence: bool
    has_comparison: bool
    has_conclusion: bool
    has_extraction_error: bool
    boilerplate_ratio: float
    table_noise_ratio: float
    front_loaded_noise: bool


def clean_text_for_excel(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    stripped = cleaned.lstrip()
    # Prevent Excel from treating source text as a formula such as "=..."
    if stripped.startswith(("=", "+", "-", "@")):
        prefix_len = len(cleaned) - len(stripped)
        cleaned = cleaned[:prefix_len] + "'" + stripped
    return cleaned


def lower_blob(*parts: str) -> str:
    return " ".join(part for part in parts if part).lower()


def contains_any(text: str, terms: tuple[str, ...]) -> bool:
    return any(term in text for term in terms)


def boilerplate_ratio(text: str) -> float:
    total = max(word_count(text), 1)
    boilerplate_words = 0
    for line in text.splitlines():
        lower = line.lower()
        if any(term in lower for term in BOILERPLATE_TERMS):
            boilerplate_words += word_count(line)
    return boilerplate_words / total


def table_noise_ratio(text: str) -> float:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return 1.0
    noisy = 0
    for line in lines:
        if re.search(r"\.{8,}", line):
            noisy += 1
            continue
        words = word_count(line)
        tabs = line.count("\t")
        digits = len(NUMBER_RE.findall(line))
        if tabs >= 5 or digits >= 20 or (digits >= max(5, words // 2) and words < 45):
            noisy += 1
    return noisy / len(lines)


def front_loaded_noise(text: str) -> bool:
    first_words = " ".join(NUMBER_RE.sub(" ", text).split()[:120]).lower()
    strong_starts = (
        "skip to main content",
        "cookies on",
        "more resources",
        "access essential accompanying documents",
        "these risks and uncertainties",
        "important factors that could cause",
        "forward-looking statements",
        "unfavorable conditions in our industry",
        "condensed consolidated statements",
    )
    return any(term in first_words for term in strong_starts)


def classify_domain(text: str, file_names: str) -> str:
    blob = lower_blob(text[:12000], file_names)
    scores: Counter[str] = Counter()
    for domain, keywords in DOMAIN_KEYWORDS.items():
        for keyword in keywords:
            scores[domain] += blob.count(keyword)
    if not scores:
        return "General policy and evidence"
    domain, score = scores.most_common(1)[0]
    return domain if score > 0 else "General policy and evidence"


def classify_source_type(files: list[dict[str, str]]) -> str:
    names = lower_blob(*(row.get("file_name", "") for row in files))
    formats = {row.get("format", "") for row in files}
    if "xlsx" in formats or "xls" in formats:
        return "structured questionnaire or dataset"
    if "regulation" in names or " act " in f" {names} " or "legislation" in names:
        return "law or regulation"
    if "annual report" in names or "financial statements" in names:
        return "corporate annual or financial report"
    if "survey" in names:
        return "survey report"
    if "factsheet" in names or "fact sheet" in names:
        return "factsheet"
    if "report" in names:
        return "analytical report"
    if "press-release" in names or "press release" in names:
        return "press release"
    return "source document"


def evaluate_flags(original_excerpt: str) -> Flags:
    lower = original_excerpt.lower()
    return Flags(
        has_summary=contains_any(lower, SUMMARY_TERMS),
        has_evidence=bool(NUMBER_RE.search(original_excerpt)),
        has_comparison=contains_any(lower, COMPARISON_TERMS)
        or len(set(YEAR_RE.findall(original_excerpt))) >= 2
        or sum(1 for term in COUNTRY_TERMS if term in lower) >= 2,
        has_conclusion=contains_any(lower, CONCLUSION_TERMS),
        has_extraction_error=bool(EXTRACTION_ERROR_RE.search(original_excerpt)),
        boilerplate_ratio=boilerplate_ratio(original_excerpt),
        table_noise_ratio=table_noise_ratio(original_excerpt),
        front_loaded_noise=front_loaded_noise(original_excerpt),
    )


def dimensions_for(flags: Flags, domain: str, original_excerpt: str) -> str:
    lower = original_excerpt.lower()
    dimensions: list[str] = []
    metric_dimensions: list[str] = []

    metric_groups = {
        "GDP growth, inflation, employment, and fiscal indicators": (
            "gdp",
            "inflation",
            "employment",
            "unemployment",
            "deficit",
            "debt",
            "budget",
            "fiscal",
        ),
        "revenue, margins, cash flow, and operating performance": (
            "revenue",
            "operating income",
            "cash flow",
            "margin",
            "profit",
            "sales",
            "earnings",
        ),
        "public trust, service performance, and institutional outcomes": (
            "trust",
            "public service",
            "institution",
            "government",
            "satisfaction",
        ),
        "health prevalence, mortality, disease burden, and risk factors": (
            "health",
            "mortality",
            "disease",
            "obesity",
            "patients",
            "risk factor",
        ),
        "energy demand, emissions, prices, and efficiency indicators": (
            "energy",
            "emissions",
            "electricity",
            "renewable",
            "efficiency",
            "demand",
        ),
        "open data maturity, portal functions, reuse, and governance capacity": (
            "open data",
            "portal",
            "reuse",
            "dataset",
            "data governance",
            "public sector data",
        ),
        "AI governance, model risk, digital market rules, and institutional controls": (
            "artificial intelligence",
            " ai ",
            "model",
            "algorithm",
            "digital market",
            "cyber",
        ),
        "population size, fertility, migration, and age structure": (
            "population",
            "fertility",
            "migration",
            "ageing",
            "age structure",
            "life expectancy",
        ),
    }
    for label, keywords in metric_groups.items():
        if any(keyword in lower for keyword in keywords):
            metric_dimensions.append(label)

    if metric_dimensions:
        dimensions.extend(metric_dimensions[:2])
    if sum(1 for term in COUNTRY_TERMS if term in lower) >= 2:
        dimensions.append("country or regional differences")
    if len(set(YEAR_RE.findall(original_excerpt))) >= 2:
        dimensions.append("year-over-year or multi-year trends")
    if any(term in lower for term in ("firm", "company", "revenue", "operating income")):
        dimensions.append("firm-level operating and financial comparisons")
    if any(term in lower for term in ("policy", "regulation", "act", "law", "directive")):
        dimensions.append("policy instruments, regulatory scope, and implementation effects")
    if any(
        term in lower
        for term in (
            "artificial intelligence",
            " ai ",
            "digital",
            "technology",
            "model risk",
            "open data",
            "data governance",
        )
    ):
        dimensions.append("technology adoption, data governance, and institutional capacity")
    if flags.has_evidence:
        dimensions.append("quantitative indicators and reported trends")
    if not dimensions:
        return ""
    return "; ".join(dict.fromkeys(dimensions))


def synthesis_for(domain: str, dimensions: str, flags: Flags) -> str:
    if not dimensions or not flags.has_evidence or not flags.has_comparison:
        return ""
    implication = "with interpretation or implications" if flags.has_conclusion else "with limited explicit implications"
    return (
        f"Provides evidence for Core tasks comparing {dimensions} in the "
        f"{domain.lower()} domain, {implication}."
    )


def theme_for(domain: str, source_type: str, dimensions: str) -> str:
    if not dimensions:
        return f"Evidence review using a {source_type}"
    first_dimension = dimensions.split("; ", 1)[0]
    return f"Comparative report on {first_dimension} using {source_type} evidence"


def load_index(index_path: Path) -> dict[str, SourceMeta]:
    rows: list[dict[str, str]] = []
    if index_path.exists():
        with index_path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

    by_source: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        by_source[row["source_id"].upper()].append(row)

    duplicate_groups: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        source_id = row["source_id"].upper()
        for key in ("duplicate_file_group", "text_sha256"):
            value = row.get(key, "")
            if value:
                duplicate_groups[f"{key}:{value}"].add(source_id)

    canonical_by_source: dict[str, tuple[str, str, list[str]]] = {}
    for reason, members in duplicate_groups.items():
        if len(members) < 2:
            continue
        ordered = sorted(members)
        canonical = ordered[0]
        for member in ordered:
            existing = canonical_by_source.get(member)
            if existing is None or canonical < existing[0]:
                canonical_by_source[member] = (canonical, reason, ordered)

    meta: dict[str, SourceMeta] = {}
    for source_id, files in by_source.items():
        canonical, reason, variants = canonical_by_source.get(
            source_id, (source_id, "", [source_id])
        )
        variant_files = [
            row.get("file_path", "")
            for row in rows
            if row["source_id"].upper() in set(variants)
        ]
        meta[source_id] = SourceMeta(
            source_id=source_id,
            files=files,
            canonical_source_id=canonical,
            variant_source_ids=[sid for sid in variants if sid != source_id],
            variant_files=variant_files,
            duplicate_reason=reason,
        )
    return meta


def status_word(ok: bool, review: bool = False) -> str:
    if ok:
        return "PASS"
    return "NEEDS_REVIEW" if review else "FAIL"


def extraction_status_for(total_words: int, flags: Flags) -> str:
    if total_words == 0 or total_words < 300 or flags.has_extraction_error:
        return "REJECTED"
    if flags.boilerplate_ratio >= 0.30 or flags.table_noise_ratio >= 0.35:
        return "REJECTED"
    if total_words < 500:
        return "PARTIAL"
    if (
        flags.boilerplate_ratio >= 0.15
        or flags.table_noise_ratio >= 0.20
        or flags.front_loaded_noise
    ):
        return "NEEDS_REVIEW"
    return "COMPLETE"


def final_decision(
    source_id: str,
    meta: SourceMeta,
    total_words: int,
    flags: Flags,
    dimensions: str,
    synthesis: str,
) -> tuple[str, bool, str]:
    reasons: list[str] = []

    if meta.canonical_source_id != source_id:
        reasons.append("duplicate_variant")
    if total_words == 0:
        reasons.append("empty_excerpt")
    elif total_words < 300:
        reasons.append("word_count_below_300")
    elif total_words < 500:
        reasons.append("word_count_300_499")
    if flags.has_extraction_error:
        reasons.append("extraction_error")
    if flags.boilerplate_ratio >= 0.30:
        reasons.append("boilerplate_dominance")
    elif flags.boilerplate_ratio >= 0.15:
        reasons.append("boilerplate_noise")
    if flags.table_noise_ratio >= 0.35:
        reasons.append("table_or_toc_noise")
    elif flags.table_noise_ratio >= 0.20:
        reasons.append("table_or_toc_noise_review")
    if flags.front_loaded_noise:
        reasons.append("front_loaded_boilerplate")
    if not flags.has_evidence:
        reasons.append("no_quantitative_evidence")
    if not flags.has_comparison:
        reasons.append("no_comparison_potential")
    if not synthesis:
        reasons.append("no_synthesis_potential")
    if not dimensions:
        reasons.append("missing_comparison_dimensions")

    hard_accept = not reasons
    if hard_accept:
        return "COMPLETE", True, ""

    reject_reasons = {
        "duplicate_variant",
        "empty_excerpt",
        "word_count_below_300",
        "extraction_error",
        "boilerplate_dominance",
        "no_quantitative_evidence",
        "table_or_toc_noise",
    }
    if any(reason in reject_reasons for reason in reasons):
        return "REJECTED", False, "; ".join(reasons)

    if "word_count_300_499" in reasons:
        return "PARTIAL", False, "; ".join(reasons)

    review_reasons = {
        "boilerplate_noise",
        "front_loaded_boilerplate",
        "table_or_toc_noise_review",
        "no_comparison_potential",
        "no_synthesis_potential",
        "missing_comparison_dimensions",
    }
    if any(reason in review_reasons for reason in reasons):
        return "NEEDS_REVIEW", False, "; ".join(reasons)

    return "NEEDS_REVIEW", False, "; ".join(reasons)


def build_row(source_id: str, raw_text: str, meta: SourceMeta, logger: logging.Logger) -> dict[str, object]:
    selection = select_excerpts(source_id, raw_text, logger)
    excerpts = [selection.excerpt_1, selection.excerpt_2, selection.excerpt_3]
    original_excerpt = "\n\n--- excerpt break ---\n\n".join(
        excerpt for excerpt in excerpts if excerpt
    )
    total_words = word_count(original_excerpt)
    flags = evaluate_flags(original_excerpt)
    file_names = "; ".join(row.get("file_name", "") for row in meta.files)
    domain = classify_domain(original_excerpt, file_names)
    source_type = classify_source_type(meta.files)
    dimensions = dimensions_for(flags, domain, original_excerpt)
    synthesis = synthesis_for(domain, dimensions, flags)
    theme = theme_for(domain, source_type, dimensions)
    extraction_status = extraction_status_for(total_words, flags)
    decision, hard_accept, reasons = final_decision(
        source_id, meta, total_words, flags, dimensions, synthesis
    )

    source_exists = bool(meta.files) and all(
        Path(row.get("file_path", "")).name.startswith(source_id)
        for row in meta.files
    )
    source_mapping_review = len(meta.files) > 1 and meta.canonical_source_id == source_id
    if decision == "COMPLETE" and source_mapping_review:
        decision = "NEEDS_REVIEW"
        hard_accept = False
        reasons = "multi_file_source_requires_review"

    summary_or_interpretation = flags.has_summary or flags.has_conclusion
    semantic_pass = (
        summary_or_interpretation
        and flags.has_evidence
        and flags.has_comparison
        and flags.boilerplate_ratio < 0.30
        and not flags.has_extraction_error
    )
    utility_pass = hard_accept or (
        total_words >= 500 and flags.has_evidence and flags.has_comparison
    )
    annotation_pass = bool(dimensions and synthesis and theme)

    return {
        "source_id": source_id,
        "canonical_source_id": meta.canonical_source_id,
        "is_canonical": meta.canonical_source_id == source_id,
        "variant_source_ids": "; ".join(meta.variant_source_ids),
        "variant_files": "; ".join(meta.variant_files),
        "duplicate_reason": meta.duplicate_reason,
        "domain": domain,
        "source_type": source_type,
        "file_names": file_names,
        "original_excerpt": original_excerpt,
        "translated_excerpt": "",
        "translation_status": "pending",
        "excerpt_1": selection.excerpt_1,
        "excerpt_2": selection.excerpt_2,
        "excerpt_3": selection.excerpt_3,
        "word_count": total_words,
        "short_excerpt": total_words < 500,
        "has_summary_or_interpretation": summary_or_interpretation,
        "has_quantitative_evidence": flags.has_evidence,
        "has_comparison_potential": flags.has_comparison,
        "has_conclusions_or_implications": flags.has_conclusion,
        "boilerplate_ratio": round(flags.boilerplate_ratio, 3),
        "table_noise_ratio": round(flags.table_noise_ratio, 3),
        "front_loaded_noise": flags.front_loaded_noise,
        "has_extraction_error": flags.has_extraction_error,
        "comparison_dimensions": dimensions,
        "synthesis_potential": synthesis,
        "likely_report_theme": theme,
        "extraction_status": extraction_status,
        "level_1_source_integrity": status_word(
            source_exists and total_words > 0 and not flags.has_extraction_error
        ),
        "level_1_source_mapping": status_word(
            source_exists, review=source_mapping_review
        ),
        "level_2_duplicate_handling": status_word(
            meta.canonical_source_id == source_id,
            review=bool(meta.duplicate_reason),
        ),
        "level_3_text_extraction_quality": status_word(
            total_words >= 500
            and not flags.has_extraction_error
            and flags.boilerplate_ratio < 0.30,
            review=300 <= total_words < 500,
        ),
        "level_4_semantic_quality": status_word(
            semantic_pass,
            review=summary_or_interpretation
            and flags.has_evidence
            and flags.has_comparison
            and not flags.has_conclusion,
        ),
        "level_5_task_utility": status_word(utility_pass),
        "level_6_translation_quality": "NEEDS_REVIEW",
        "level_7_annotation_utility": status_word(annotation_pass),
        "hard_accept": hard_accept,
        "final_decision": decision,
        "rejection_reasons": reasons,
    }


def write_packet(rows: list[dict[str, object]], output_path: Path) -> None:
    headers = [
        "source_id",
        "canonical_source_id",
        "is_canonical",
        "variant_source_ids",
        "variant_files",
        "duplicate_reason",
        "domain",
        "source_type",
        "file_names",
        "original_excerpt",
        "translated_excerpt",
        "translation_status",
        "excerpt_1",
        "excerpt_2",
        "excerpt_3",
        "word_count",
        "short_excerpt",
        "has_summary_or_interpretation",
        "has_quantitative_evidence",
        "has_comparison_potential",
        "has_conclusions_or_implications",
        "boilerplate_ratio",
        "table_noise_ratio",
        "front_loaded_noise",
        "has_extraction_error",
        "comparison_dimensions",
        "synthesis_potential",
        "likely_report_theme",
        "extraction_status",
        "level_1_source_integrity",
        "level_1_source_mapping",
        "level_2_duplicate_handling",
        "level_3_text_extraction_quality",
        "level_4_semantic_quality",
        "level_5_task_utility",
        "level_6_translation_quality",
        "level_7_annotation_utility",
        "hard_accept",
        "final_decision",
        "rejection_reasons",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "core_extraction"
    ws.append(headers)

    fills = {
        "COMPLETE": PatternFill("solid", fgColor="D9EAD3"),
        "PARTIAL": PatternFill("solid", fgColor="FFF2CC"),
        "NEEDS_REVIEW": PatternFill("solid", fgColor="FCE5CD"),
        "REJECTED": PatternFill("solid", fgColor="F4CCCC"),
    }

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = clean_text_for_excel(row.get(header, ""))
            if isinstance(cell.value, str):
                cell.data_type = "s"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        decision = str(row.get("final_decision", ""))
        if decision in fills:
            for cell in ws[row_idx]:
                cell.fill = fills[decision]

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for idx, header in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        if header in {"original_excerpt", "excerpt_1", "excerpt_2", "excerpt_3"}:
            width = 80
        elif header in {
            "variant_files",
            "comparison_dimensions",
            "synthesis_potential",
            "likely_report_theme",
            "rejection_reasons",
            "file_names",
        }:
            width = 38
        else:
            width = 18
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def run(raw_text_dir: Path, index_path: Path, output_path: Path, log_path: Path) -> int:
    logger = setup_logging(log_path)
    source_meta = load_index(index_path)
    raw_paths = sorted(
        path for path in raw_text_dir.glob("*.txt") if SOURCE_ID_RE.match(path.name)
    )
    if not raw_paths:
        logger.error("no raw_text source files found in %s", raw_text_dir)
        return 1

    rows: list[dict[str, object]] = []
    for path in raw_paths:
        source_id = path.stem.upper()
        try:
            raw_text = path.read_text(encoding="utf-8")
            meta = source_meta.get(
                source_id,
                SourceMeta(source_id, [], source_id, [], [], ""),
            )
            rows.append(build_row(source_id, raw_text, meta, logger))
            logger.info("built QC row for %s", source_id)
        except Exception as exc:
            logger.exception("failed to build row for %s: %s", source_id, exc)
            rows.append(
                {
                    "source_id": source_id,
                    "canonical_source_id": source_id,
                    "final_decision": "REJECTED",
                    "hard_accept": False,
                    "rejection_reasons": f"row_build_failed: {exc}",
                    "level_1_source_integrity": "FAIL",
                }
            )

    write_packet(rows, output_path)
    decisions = Counter(str(row.get("final_decision", "")) for row in rows)
    logger.info("wrote %s (%d rows)", output_path, len(rows))
    logger.info("decision counts: %s", dict(decisions))
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    base_dir = Path(__file__).resolve().parent
    parser.add_argument("--raw-text-dir", type=Path, default=base_dir / "raw_text")
    parser.add_argument("--index", type=Path, default=base_dir / "raw_sources_index.csv")
    parser.add_argument("--output", type=Path, default=base_dir / "core_extraction.xlsx")
    parser.add_argument("--log", type=Path, default=base_dir / "core_extraction_qc.log")
    args = parser.parse_args()
    raise SystemExit(run(args.raw_text_dir, args.index, args.output, args.log))


if __name__ == "__main__":
    main()
