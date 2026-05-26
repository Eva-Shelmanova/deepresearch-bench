#!/usr/bin/env python3
"""Extract bounded entity lists from downloaded Set source files in set/."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None  # type: ignore

BASE = Path("/Users/evashelmanova/Documents/IAI MSU/materials/sources_extracted")
SET_DIR = BASE / "set"
OUT_DIR = SET_DIR / "extractions"
STEP3 = BASE / "benchmark_step3_source_clustering.xlsx"
INDEX_XLSX = SET_DIR / "set_extraction.xlsx"
SUMMARY_MD = SET_DIR / "SET_EXTRACTION_SUMMARY.md"

SET_ID_RE = re.compile(r"^(SET\d{3})_", re.I)
SLUG_RE = re.compile(r"[^A-Za-z0-9]+")

SKIP_LINK_PREFIXES = (
    "Special:", "Help:", "Wikipedia:", "File:", "Category:", "Portal:",
    "Template:", "Talk:", "User:", "Draft:", "#",
)

NOISE_LABELS = {
    "edit", "show", "hide", "view", "talk", "main article", "main page",
    "coordinates", "references", "external links", "see also", "notes",
    "bibliography", "further reading", "navigation menu", "search",
}


def slugify(text: str, max_len: int = 80) -> str:
    s = SLUG_RE.sub("_", text).strip("_")
    return s[:max_len] or "untitled"


def load_step3_metadata() -> dict[str, dict]:
    if not STEP3.exists():
        return {}
    wb = openpyxl.load_workbook(STEP3, read_only=True, data_only=True)
    ws = wb["Set Clusters"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    meta: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        sid = str(d.get("Source ID") or "").strip()
        if sid and sid not in meta:
            meta[sid] = d
    wb.close()
    return meta


def classify_page(title: str, filename: str) -> str:
    t = f"{title} {filename}".lower()
    if "list of world heritage sites in " in t:
        return "country_wh_site_list"
    if "lists of world heritage" in t or "list of unesco world heritage" in t:
        return "wh_index"
    if "list of airports in " in t:
        return "country_airport_list"
    if "list of airports by iata" in t:
        return "airport_iata_index"
    if re.search(r"list of countries|list of sovereign|national capitals|largest cities", t):
        return "country_data_table"
    if re.search(r"list of .* by ", t):
        return "generic_list_table"
    if "comparison of" in t and "framework" in t:
        return "comparison_table"
    if " - wikipedia" in filename.lower() or filename.lower().endswith(".htm"):
        return "wikipedia_article"
    if filename.lower().endswith(".pdf"):
        return "pdf_document"
    return "unknown"


def normalize_href(href: str) -> str:
    href = (href or "").strip()
    if href.startswith("//"):
        href = "https:" + href
    if href.startswith("/wiki/"):
        href = "https://en.wikipedia.org" + href
    return href.split("#")[0]


def is_noise_label(label: str) -> bool:
    label = re.sub(r"\s+", " ", label or "").strip()
    if not label or len(label) < 2:
        return True
    lower = label.lower()
    if lower in NOISE_LABELS:
        return True
    if lower.startswith("[") and lower.endswith("]"):
        return True
    if re.fullmatch(r"\[\d+\]", label):
        return True
    return False


def extract_wikipedia_html(path: Path) -> dict[str, Any]:
    html = path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(html, "lxml")

    title_el = soup.select_one("#firstHeading") or soup.find("h1")
    title = title_el.get_text(" ", strip=True) if title_el else path.stem
    content = soup.select_one("#mw-content-text .mw-parser-output") or soup.select_one("#mw-content-text")

    lead: list[str] = []
    if content:
        for p in content.find_all("p", recursive=False):
            txt = p.get_text(" ", strip=True)
            if len(txt) > 40:
                lead.append(txt)
            if len(lead) >= 2:
                break

    page_class = classify_page(title, path.name)
    table_entities: list[dict] = []
    link_entities: list[dict] = []

    if content:
        for ti, table in enumerate(content.select("table.wikitable")):
            caption = ""
            cap = table.find("caption")
            if cap:
                caption = cap.get_text(" ", strip=True)
            for ri, tr in enumerate(table.select("tr")):
                cells = [c.get_text(" ", strip=True) for c in tr.select("th,td")]
                cells = [c for c in cells if c]
                if len(cells) < 1:
                    continue
                label = cells[0]
                if is_noise_label(label) or label.lower() in {"name", "country", "rank", "#"}:
                    continue
                if len(label) > 250:
                    continue
                table_entities.append(
                    {
                        "label": label,
                        "cells": cells,
                        "table_index": ti,
                        "row_index": ri,
                        "caption": caption,
                        "entity_type": "wikitable_row",
                    }
                )

        seen_urls: set[str] = set()
        for a in content.find_all("a", href=True):
            href = normalize_href(a["href"])
            if not href or "wikipedia.org/wiki/" not in href:
                continue
            if any(href.split("/wiki/")[-1].startswith(p) for p in SKIP_LINK_PREFIXES):
                continue
            if href in seen_urls:
                continue
            label = a.get_text(" ", strip=True)
            if is_noise_label(label):
                continue
            seen_urls.add(href)
            entity_type = "wiki_link"
            is_wh_country = "List_of_World_Heritage_Sites_in_" in href
            is_airport_country = "List_of_airports_in_" in href
            if is_wh_country:
                entity_type = "country_or_region_world_heritage_list"
            elif is_airport_country:
                entity_type = "country_airport_list_page"
            link_entities.append(
                {
                    "label": label,
                    "url": href,
                    "entity_type": entity_type,
                    "is_country_or_region_list": is_wh_country,
                }
            )

    # Deduplicate table entities by label
    deduped_tables: list[dict] = []
    seen_labels: set[str] = set()
    for ent in table_entities:
        key = ent["label"].lower()
        if key in seen_labels:
            continue
        seen_labels.add(key)
        deduped_tables.append(ent)

    entities = deduped_tables if len(deduped_tables) >= 3 else []
    extraction_mode = "wikitable"

    if page_class in {"wh_index", "airport_iata_index"} and link_entities:
        entities = link_entities
        extraction_mode = "wiki_links"
    elif page_class == "country_wh_site_list" and deduped_tables:
        entities = deduped_tables
        extraction_mode = "wikitable_sites"
    elif page_class == "country_airport_list" and deduped_tables:
        entities = deduped_tables
        extraction_mode = "wikitable_airports"
    elif page_class in {"country_data_table", "generic_list_table", "comparison_table"} and deduped_tables:
        entities = deduped_tables
        extraction_mode = "wikitable"
    elif not entities and link_entities:
        entities = link_entities[:500]
        extraction_mode = "wiki_links_fallback"
    elif page_class == "wikipedia_article" and title:
        entities = [{"label": title, "entity_type": "single_article", "url": ""}]
        extraction_mode = "single_article"

    return {
        "source_file": path.name,
        "page_title": title,
        "page_class": page_class,
        "extraction_mode": extraction_mode,
        "lead_paragraphs": lead,
        "entities": entities,
        "counts": {
            "entities_extracted": len(entities),
            "wikitable_rows": len(deduped_tables),
            "wiki_links": len(link_entities),
        },
    }


def extract_pdf(path: Path) -> dict[str, Any]:
    if PdfReader is None:
        return {
            "source_file": path.name,
            "page_title": path.stem,
            "page_class": "pdf_document",
            "extraction_mode": "pdf_unavailable",
            "entities": [],
            "error": "pypdf not installed",
            "counts": {"entities_extracted": 0},
        }
    try:
        reader = PdfReader(str(path))
        pages_text = []
        for page in reader.pages[:30]:
            pages_text.append(page.extract_text() or "")
        text = "\n".join(pages_text)
        lines = [ln.strip() for ln in text.splitlines() if 3 < len(ln.strip()) < 200]
        entities = [
            {"label": ln, "entity_type": "pdf_line", "line_index": i}
            for i, ln in enumerate(lines[:200])
            if not re.match(r"^\d+$", ln)
        ]
        return {
            "source_file": path.name,
            "page_title": path.stem,
            "page_class": "pdf_document",
            "extraction_mode": "pdf_text_lines",
            "lead_paragraphs": [text[:1500]] if text else [],
            "entities": entities,
            "counts": {
                "entities_extracted": len(entities),
                "pdf_pages_read": min(len(reader.pages), 30),
            },
        }
    except Exception as exc:
        return {
            "source_file": path.name,
            "page_title": path.stem,
            "page_class": "pdf_document",
            "extraction_mode": "pdf_error",
            "entities": [],
            "error": str(exc),
            "counts": {"entities_extracted": 0},
        }


def infer_set_annotation(page_class: str, entity_count: int) -> dict[str, str]:
    if page_class == "wh_index":
        return {
            "candidate_answer_set_type": "linked_country_or_region_world_heritage_list_pages",
            "boundedness": "bounded" if entity_count >= 50 else "partial",
            "qa_status": "pass_as_index_source" if entity_count >= 50 else "needs_review",
        }
    if page_class == "country_wh_site_list":
        return {
            "candidate_answer_set_type": "individual_world_heritage_sites_in_country",
            "boundedness": "bounded" if entity_count >= 1 else "empty",
            "qa_status": "pass" if entity_count >= 1 else "needs_review",
        }
    if page_class == "country_airport_list":
        return {
            "candidate_answer_set_type": "airports_in_country",
            "boundedness": "bounded" if entity_count >= 1 else "empty",
            "qa_status": "pass" if entity_count >= 1 else "needs_review",
        }
    if page_class in {"country_data_table", "generic_list_table"}:
        return {
            "candidate_answer_set_type": "wikitable_entities",
            "boundedness": "bounded" if entity_count >= 10 else "partial",
            "qa_status": "pass" if entity_count >= 10 else "needs_review",
        }
    return {
        "candidate_answer_set_type": "unknown",
        "boundedness": "unknown",
        "qa_status": "needs_review" if entity_count < 3 else "pass",
    }


def discover_source_files() -> list[Path]:
    files: list[Path] = []
    for path in sorted(SET_DIR.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".htm", ".html", ".pdf"}:
            continue
        if SET_ID_RE.match(path.name):
            files.append(path)
    return files


def write_index_xlsx(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "set_extraction"
    headers = [
        "source_id", "source_file", "page_title", "page_class", "extraction_mode",
        "entity_count", "qa_status", "boundedness", "answer_set_type",
        "cluster_id", "domain", "url_step3", "extraction_json",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="FF1F3864")
    font = Font(color="FFFFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    widths = [10, 55, 45, 22, 20, 12, 14, 12, 35, 14, 28, 50, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(INDEX_XLSX)


def write_summary(rows: list[dict], by_set: dict[str, list[dict]]) -> None:
    lines = [
        "# Set Extraction Summary",
        "",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
        f"- Source files processed: **{len(rows)}**",
        f"- Unique SET IDs: **{len(by_set)}**",
        f"- Output directory: `set/extractions/`",
        f"- Index workbook: `set/set_extraction.xlsx`",
        "",
        "## By page class",
        "",
    ]
    for cls, n in Counter(r["page_class"] for r in rows).most_common():
        lines.append(f"- {cls}: {n} files")
    lines.extend(["", "## By SET ID (file counts)", ""])
    for sid in sorted(by_set):
        cnt = len(by_set[sid])
        entities = sum(r["entity_count"] for r in by_set[sid])
        lines.append(f"- **{sid}**: {cnt} files, {entities} entities total")
    lines.extend(["", "## Notes", ""])
    lines.append("- `wh_index` / `country_wh_site_list` / `country_airport_list` are Wikipedia list pages.")
    lines.append("- Multi-file SET IDs (e.g. SET008, SET009) mean you downloaded sub-pages; each file has its own JSON.")
    lines.append("- PDF rows are line-based placeholders; refine manually for registry PDFs.")
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    meta = load_step3_metadata()
    files = discover_source_files()
    print(f"Found {len(files)} Set source files")

    index_rows: list[dict] = []
    by_set: dict[str, list[dict]] = defaultdict(list)

    for i, path in enumerate(files, start=1):
        sid = SET_ID_RE.match(path.name).group(1).upper()
        m = meta.get(sid, {})
        print(f"[{i}/{len(files)}] {path.name[:70]}")

        if path.suffix.lower() in {".htm", ".html"}:
            payload = extract_wikipedia_html(path)
        else:
            payload = extract_pdf(path)

        ann = infer_set_annotation(payload["page_class"], len(payload.get("entities", [])))
        payload["source_id"] = sid
        payload["extracted_at_utc"] = datetime.now(timezone.utc).isoformat()
        payload["set_annotation"] = ann
        if m:
            payload["step3_metadata"] = {
                "domain": m.get("Domain"),
                "source_type": m.get("Source type"),
                "cluster_id": m.get("Cluster ID"),
                "url": m.get("URL"),
            }

        out_sub = OUT_DIR / sid
        out_sub.mkdir(parents=True, exist_ok=True)
        out_name = slugify(path.stem.replace(sid + "_", "", 1)) + ".json"
        out_path = out_sub / out_name
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        row = {
            "source_id": sid,
            "source_file": path.name,
            "page_title": payload.get("page_title", ""),
            "page_class": payload.get("page_class", ""),
            "extraction_mode": payload.get("extraction_mode", ""),
            "entity_count": payload.get("counts", {}).get("entities_extracted", 0),
            "qa_status": ann.get("qa_status", ""),
            "boundedness": ann.get("boundedness", ""),
            "answer_set_type": ann.get("candidate_answer_set_type", ""),
            "cluster_id": m.get("Cluster ID", ""),
            "domain": m.get("Domain", ""),
            "url_step3": m.get("URL", ""),
            "extraction_json": str(out_path.relative_to(BASE)),
        }
        index_rows.append(row)
        by_set[sid].append(row)

    write_index_xlsx(index_rows)
    write_summary(index_rows, by_set)
    print(f"Wrote {INDEX_XLSX}")
    print(f"Wrote {SUMMARY_MD}")
    print(f"JSON files in {OUT_DIR}")


if __name__ == "__main__":
    main()
