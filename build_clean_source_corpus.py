#!/usr/bin/env python3
"""Build one clean, source-divided Core + Set corpus file."""

from __future__ import annotations

import csv
import html
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from bs4 import BeautifulSoup

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None  # type: ignore


BASE = Path("/Users/evashelmanova/Documents/IAI MSU/materials/sources_extracted")
RAW_TEXT_DIR = BASE / "raw_text"
RAW_TEXT_SET_DIR = BASE / "raw_text_set"
CORE_XLSX = BASE / "core_extraction.xlsx"
RAW_INDEX = BASE / "raw_sources_index.csv"
RAW_SET_INDEX = BASE / "raw_set_sources_index.csv"
SET_DIR = BASE / "set"
SET_INDEX_XLSX = SET_DIR / "set_extraction.xlsx"
OUT_PATH = BASE / "clean_core_set_sources.md"

SET_FILE_RE = re.compile(r"^(SET\d{3})_", re.I)
SPACE_RE = re.compile(r"[ \t\u00a0]+")
PAGE_MARKER_RE = re.compile(
    r"^(?:[-–—\s]*)?(?:page\s*)?\d{1,4}(?:\s*(?:of|/)\s*\d{1,4})?(?:[-–—\s]*)?$",
    re.I,
)
PDF_FOOTER_RE = re.compile(r"^[\W_]*\d{1,4}\s*[|]\s*.*$|^.*[|]\s*\d{1,4}[\W_]*$")

HTML_REMOVE_SELECTORS = [
    "script",
    "style",
    "noscript",
    "meta",
    "link",
    "svg",
    "figure",
    "iframe",
    "sup.reference",
    ".mw-editsection",
    ".navbox",
    ".metadata",
    ".ambox",
    ".hatnote",
    ".mw-empty-elt",
    "#toc",
    ".toc",
    "header",
    "footer",
    "nav",
]


def normalize_spaces(text: str) -> str:
    text = text.replace("\uf0fc", "-").replace("\uf0b7", "-").replace("\uf07c", "|")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(SPACE_RE.sub(" ", line).strip() for line in text.splitlines())


def is_page_noise(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if PAGE_MARKER_RE.fullmatch(s) and len(s) <= 20:
        return True
    if PDF_FOOTER_RE.fullmatch(s) and len(s) <= 180:
        return True
    if re.fullmatch(r"--\s*\d{1,4}\s+of\s+\d{1,4}\s*--", s, re.I):
        return True
    if re.fullmatch(r"\[\s*page\s+\d{1,4}\s*\]", s, re.I):
        return True
    return False


def repeated_noise_lines(lines: list[str]) -> set[str]:
    normalized = [line.strip() for line in lines if line.strip()]
    counts = Counter(normalized)
    noisy: set[str] = set()
    markers = (
        "©",
        "copyright",
        "all rights reserved",
        "isbn",
        "doi:",
        "www.",
        "http://",
        "https://",
    )
    for line, count in counts.items():
        lower = line.lower()
        if count >= 5 and len(line) <= 180 and any(marker in lower for marker in markers):
            noisy.add(line)
        elif count >= 8 and len(line) <= 140 and line.isupper():
            noisy.add(line)
    return noisy


def clean_text(text: str) -> str:
    text = normalize_spaces(text)
    raw_lines = text.splitlines()
    repeated_noise = repeated_noise_lines(raw_lines)

    lines: list[str] = []
    previous_blank = False
    for raw in raw_lines:
        line = raw.strip()
        if line in repeated_noise:
            continue
        if is_page_noise(line):
            continue
        if not line:
            if not previous_blank:
                lines.append("")
            previous_blank = True
            continue
        lines.append(line)
        previous_blank = False

    return paragraphize(lines).strip()


def looks_like_table_line(line: str) -> bool:
    if "\t" in line or "|" in line:
        return True
    if len(re.findall(r"\s{2,}", line)) >= 2:
        return True
    return False


def looks_like_heading(line: str) -> bool:
    if len(line) > 120:
        return False
    if line.endswith(":"):
        return True
    letters = [c for c in line if c.isalpha()]
    if letters and sum(c.isupper() for c in letters) / len(letters) > 0.75:
        return True
    return False


def paragraphize(lines: list[str]) -> str:
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        if not line:
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(line)
    if current:
        blocks.append(current)

    rendered: list[str] = []
    for block in blocks:
        if any(looks_like_table_line(line) for line in block):
            rendered.append("\n".join(block))
        elif len(block) == 1:
            rendered.append(block[0])
        elif all(looks_like_heading(line) for line in block[:2]) and len(block) <= 3:
            rendered.append("\n".join(block))
        else:
            rendered.append(" ".join(block))
    return "\n\n".join(rendered)


def read_core_metadata() -> dict[str, dict[str, Any]]:
    if not CORE_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(CORE_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: dict[str, dict[str, Any]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        source_id = str(row.get("source_id") or row.get("Source ID") or "").strip()
        if source_id:
            rows[source_id] = row
    wb.close()
    return rows


def read_raw_index() -> dict[str, list[dict[str, str]]]:
    if not RAW_INDEX.exists():
        return {}
    by_source: dict[str, list[dict[str, str]]] = {}
    with RAW_INDEX.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            sid = (row.get("source_id") or "").strip()
            if sid:
                by_source.setdefault(sid, []).append(row)
    return by_source


def read_raw_set_index() -> dict[str, dict[str, str]]:
    """Map set/ file_name -> raw_set_sources_index row."""
    if not RAW_SET_INDEX.exists():
        return {}
    by_name: dict[str, dict[str, str]] = {}
    with RAW_SET_INDEX.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            name = (row.get("file_name") or "").strip()
            if name:
                by_name[name] = row
    return by_name


def read_set_index() -> list[dict[str, Any]]:
    if not SET_INDEX_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(SET_INDEX_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(headers, values)) for values in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return rows


def md_escape_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("\n", " ").replace("|", "\\|").strip()


def metadata_lines(items: list[tuple[str, Any]]) -> list[str]:
    out: list[str] = []
    for label, value in items:
        if value is None or value == "":
            continue
        out.append(f"- **{label}:** {value}")
    return out


def fenced_text(text: str) -> str:
    if "```" not in text:
        return f"```text\n{text}\n```"
    return f"````text\n{text}\n````"


def html_visible_text(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(raw, "lxml")
    for selector in HTML_REMOVE_SELECTORS:
        for tag in soup.select(selector):
            tag.decompose()

    root = (
        soup.select_one("#mw-content-text .mw-parser-output")
        or soup.select_one("#mw-content-text")
        or soup.select_one("main")
        or soup.body
        or soup
    )

    parts: list[str] = []
    for tag in root.find_all(["h1", "h2", "h3", "h4", "p", "li", "table"], recursive=True):
        name = tag.name.lower()
        if name.startswith("h"):
            text = tag.get_text(" ", strip=True)
            if text:
                parts.append("")
                parts.append(text)
        elif name in {"p", "li"}:
            text = tag.get_text(" ", strip=True)
            if len(text) >= 3:
                parts.append(text)
        elif name == "table":
            rows: list[str] = []
            for tr in tag.select("tr"):
                cells = [cell.get_text(" ", strip=True) for cell in tr.select("th,td")]
                cells = [cell for cell in cells if cell]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                parts.append("")
                parts.extend(rows)

    if not parts:
        parts = [root.get_text("\n", strip=True)]
    return clean_text("\n".join(parts))


def pdf_text(path: Path) -> str:
    if PdfReader is None:
        return ""
    reader = PdfReader(str(path))
    return clean_text("\n\n".join(page.extract_text() or "" for page in reader.pages))


def set_source_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".html", ".htm"}:
        return html_visible_text(path)
    if suffix == ".pdf":
        return pdf_text(path)
    return ""


def render_set_entities(payload: dict[str, Any]) -> str:
    entities = payload.get("entities") or []
    if not entities:
        return ""

    lines: list[str] = []
    current_caption: str | None = None
    for ent in entities:
        cells = ent.get("cells")
        caption = ent.get("caption") or ""
        if cells:
            if caption and caption != current_caption:
                lines.append("")
                lines.append(f"Table: {caption}")
                current_caption = caption
            lines.append(" | ".join(md_escape_cell(cell) for cell in cells))
        else:
            label = md_escape_cell(ent.get("label"))
            url = md_escape_cell(ent.get("url"))
            entity_type = md_escape_cell(ent.get("entity_type"))
            if url:
                lines.append(f"{label} | {entity_type} | {url}")
            else:
                lines.append(f"{label} | {entity_type}")
    return "\n".join(line for line in lines if line is not None).strip()


def render_core_section(lines: list[str]) -> tuple[int, int]:
    metadata = read_core_metadata()
    raw_index = read_raw_index()
    source_files = sorted(RAW_TEXT_DIR.glob("S*.txt"))

    lines.extend(
        [
            "# Clean Core and Set Source Corpus",
            "",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            "",
            "This file consolidates the available Core and Set source text into one source-divided corpus.",
            "Core text is taken from `raw_text/`; Set text is taken from `raw_text_set/` (one file per downloaded Set source).",
            "The cleaning is conservative: page-number markers, repeated short copyright/footer lines, script/navigation noise, and blank-line clutter are removed, while substantive paragraphs, tables, lists, and registry rows are retained.",
            "",
            "# Core Sources",
            "",
        ]
    )

    for path in source_files:
        sid = path.stem
        row = metadata.get(sid, {})
        records = raw_index.get(sid, [])
        file_names = "; ".join(r.get("file_name", "") for r in records if r.get("file_name"))
        raw_text = path.read_text(encoding="utf-8", errors="replace")
        text = clean_text(raw_text)

        lines.extend(
            [
                f"## Core Source {sid}",
                "",
                *metadata_lines(
                    [
                        ("Source files", file_names),
                        ("Final decision", row.get("final_decision")),
                        ("Canonical source ID", row.get("canonical_source_id")),
                        ("Domain", row.get("domain")),
                        ("Source type", row.get("source_type")),
                        ("Likely report theme", row.get("likely_report_theme")),
                    ]
                ),
                "",
                "### Clean Text",
                "",
                fenced_text(text),
                "",
            ]
        )
    return len(source_files), sum(len((RAW_TEXT_DIR / f"S{i:03d}.txt").read_text(encoding="utf-8", errors="replace")) for i in range(1, 264) if (RAW_TEXT_DIR / f"S{i:03d}.txt").exists())


def render_set_section(lines: list[str]) -> int:
    rows = read_set_index()
    raw_set_by_file = read_raw_set_index()
    lines.extend(["# Set Sources", ""])

    for row in rows:
        sid = str(row.get("source_id") or "").strip()
        source_file = str(row.get("source_file") or "").strip()
        json_rel = str(row.get("extraction_json") or "").strip()
        json_path = BASE / json_rel

        payload: dict[str, Any] = {}
        if json_path.exists():
            payload = json.loads(json_path.read_text(encoding="utf-8"))

        structured = render_set_entities(payload)
        lead = "\n\n".join(payload.get("lead_paragraphs") or [])
        lead = clean_text(lead) if lead else ""

        raw_row = raw_set_by_file.get(source_file, {})
        raw_rel = (raw_row.get("raw_text_output") or "").strip()
        raw_path = BASE / raw_rel if raw_rel else None
        if raw_path and raw_path.exists():
            clean_visible = clean_text(raw_path.read_text(encoding="utf-8", errors="replace"))
        else:
            source_path = SET_DIR / source_file
            clean_visible = (
                clean_text(set_source_text(source_path)) if source_path.exists() else ""
            )

        lines.extend(
            [
                f"## Set Source {sid}: {source_file}",
                "",
                *metadata_lines(
                    [
                        ("Raw text file", raw_rel or "(not extracted)"),
                        ("Page title", row.get("page_title")),
                        ("Page class", row.get("page_class")),
                        ("Extraction mode", row.get("extraction_mode")),
                        ("Entity count", row.get("entity_count")),
                        ("QA status", row.get("qa_status")),
                        ("Boundedness", row.get("boundedness")),
                        ("Answer set type", row.get("answer_set_type")),
                        ("Cluster ID", row.get("cluster_id")),
                        ("Domain", row.get("domain")),
                        ("URL", row.get("url_step3")),
                    ]
                ),
                "",
            ]
        )
        if lead:
            lines.extend(["### Lead / Context", "", lead, ""])
        if structured:
            lines.extend(["### Structured Rows / Entities", "", fenced_text(structured), ""])
        if clean_visible:
            lines.extend(["### Clean Text", "", fenced_text(clean_visible), ""])
        else:
            lines.extend(["### Clean Text", "", "_No clean text was available for this file._", ""])
    return len(rows)


def main() -> None:
    lines: list[str] = []
    core_count, _ = render_core_section(lines)
    set_count = render_set_section(lines)
    OUT_PATH.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    size_mb = OUT_PATH.stat().st_size / (1024 * 1024)
    print(f"Wrote {OUT_PATH}")
    print(f"Core sources: {core_count}")
    print(f"Set source files: {set_count}")
    print(f"Size: {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
