#!/usr/bin/env python3
"""Add a sampled QA sheet to core_clusters.xlsx without changing clusters."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


QA_ROWS = [
    {
        "cluster_id": "C001",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Coherent AI/digital governance cluster mixing analytical, source, and structured evidence. "
            "Dimensions support cross-source synthesis on open data maturity, AI governance, policy instruments, and institutional capacity."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C010",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Coherent AI/digital governance cluster. Sources combine factsheet, source document, and structured questionnaire evidence "
            "around open data maturity, public-sector data governance, and implementation capacity."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C030",
        "qa_status": "needs_review",
        "issue_type": "dimension_drift",
        "issue_notes": (
            "Economic cluster is usable, but the shared dimensions include health prevalence and disease-burden language. "
            "This may reflect overly broad source-level dimension extraction rather than a clean economic synthesis target."
        ),
        "recommended_action": "Review source excerpts before prompt writing; narrow synthesis to macroeconomic/development indicators or recluster if health content dominates.",
    },
    {
        "cluster_id": "C041",
        "qa_status": "needs_review",
        "issue_type": "dimension_drift",
        "issue_notes": (
            "Labor/education cluster includes AI governance/model-risk dimensions that may not belong to the main labor and skills theme. "
            "The cluster may still work if the intended synthesis is skills and institutional adaptation to digital change."
        ),
        "recommended_action": "Manual review; keep only if the prompt can explicitly compare skills, labor, and digital-governance implications.",
    },
    {
        "cluster_id": "C050",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Two-source labor/education cluster has sufficient comparison dimensions around regional differences, time trends, policy instruments, and quantitative indicators."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C051",
        "qa_status": "needs_review",
        "issue_type": "too_broad",
        "issue_notes": (
            "Public governance cluster mixes GDP/fiscal, energy, and public-trust dimensions. It may be topically coherent only at a broad governance level, "
            "which risks a synthesis prompt that is too diffuse."
        ),
        "recommended_action": "Review source excerpts; narrow the cluster to governance/institutional performance or split in a later clustering revision.",
    },
    {
        "cluster_id": "C059",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Two-source public governance cluster is coherent enough: institutional outcomes, multi-year trends, policy instruments, and data/governance capacity align."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C060",
        "qa_status": "needs_review",
        "issue_type": "too_broad",
        "issue_notes": (
            "Energy/climate cluster combines macroeconomic, corporate-performance, policy, and technology dimensions. "
            "It may support a cross-sector transition prompt, but it is broader than a clean energy evidence synthesis."
        ),
        "recommended_action": "Manual review; narrow synthesis to energy transition economics or recluster if firm-finance content dominates.",
    },
    {
        "cluster_id": "C063",
        "qa_status": "needs_review",
        "issue_type": "dimension_drift",
        "issue_notes": (
            "Energy/climate pair is plausible, but includes firm-level operating and financial comparisons alongside energy/emissions indicators. "
            "This may be complementary or may indicate a mismatched source."
        ),
        "recommended_action": "Inspect both source excerpts before prompt writing; keep if finance content explains energy-transition impacts.",
    },
    {
        "cluster_id": "C064",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Energy/climate pair has coherent trend, regional, policy, and quantitative dimensions. Public-service language is secondary but not disqualifying."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C065",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Corporate performance cluster is coherent around revenue, margins, cash flow, operating performance, technology capacity, and year-over-year comparison."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C069",
        "qa_status": "needs_review",
        "issue_type": "too_broad",
        "issue_notes": (
            "Health/social cluster includes GDP, public trust, firm-level finance, and policy dimensions but lacks a clear health-specific comparison dimension in the cluster fields. "
            "Risk of topically broad rather than semantically focused synthesis."
        ),
        "recommended_action": "Review source excerpts; recluster or narrow to social outcomes if health evidence is not central.",
    },
    {
        "cluster_id": "C071",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Health/social pair is coherent: disease burden, mortality/risk factors, time trends, policy implications, and quantitative indicators are present."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C072",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Demographics pair is coherent enough for Core synthesis around population structure, migration/fertility, macroeconomic indicators, and regional trends."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
    {
        "cluster_id": "C073",
        "qa_status": "pass",
        "issue_type": "none",
        "issue_notes": (
            "Demographics pair can support synthesis on population trends, age structure, fertility/mortality, and related health-risk indicators."
        ),
        "recommended_action": "Keep cluster unchanged.",
    },
]


def run(workbook_path: Path) -> int:
    wb = load_workbook(workbook_path)
    if "cluster_qa" in wb.sheetnames:
        del wb["cluster_qa"]
    ws = wb.create_sheet("cluster_qa")

    headers = [
        "cluster_id",
        "qa_status",
        "issue_type",
        "issue_notes",
        "recommended_action",
    ]
    ws.append(headers)
    for row in QA_ROWS:
        ws.append([row[header] for header in headers])

    fills = {
        "pass": PatternFill("solid", fgColor="D9EAD3"),
        "needs_review": PatternFill("solid", fgColor="FCE5CD"),
        "fail": PatternFill("solid", fgColor="F4CCCC"),
    }

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        fill = fills.get(str(row[1].value), PatternFill())
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill

    widths = {
        "A": 14,
        "B": 16,
        "C": 20,
        "D": 90,
        "E": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(workbook_path)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path(__file__).resolve().parent / "core_clusters.xlsx",
    )
    args = parser.parse_args()
    raise SystemExit(run(args.workbook))


if __name__ == "__main__":
    main()
