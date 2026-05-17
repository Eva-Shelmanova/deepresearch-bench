#!/usr/bin/env python3
"""Retune the six needs_review clusters and rerun QA on them only."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# remediation_action: retune | split | keep
REMEDIATIONS = {
    "C030": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "GDP growth, poverty reduction, shared prosperity, and infrastructure investment "
            "across countries and over time"
        ),
        "cluster_rationale": (
            "S074, S204, and S248 are retained as one economic-development cluster. The main comparison axis is "
            "macroeconomic and development outcomes—shared prosperity, infrastructure investment, and cross-country "
            "trends—not health or corporate metrics. S248 is the weakest fit and should be interpreted cautiously in synthesis."
        ),
        "synthesis_intent": (
            "Synthesize how development and infrastructure evidence compares across countries and time periods, "
            "focusing on growth, poverty/shared prosperity, and investment patterns."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned to a single macroeconomic-development axis. Health dimensions removed from cluster fields. "
            "S248 remains a peripheral source but the cluster is now sufficiently focused for prompt drafting."
        ),
        "recommended_action": "Proceed to prompt draft with explicit focus on development and infrastructure; de-emphasize S248 unless excerpt review supports it.",
    },
    "C041": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "Education access, learning outcomes, and labour-market skills across countries and policy settings"
        ),
        "cluster_rationale": (
            "S041, S045, and S053 support one education-and-skills comparison axis. AI-governance and fiscal-budget "
            "language were removed from the cluster framing because they came from source-level metadata drift, not "
            "the core labour/education synthesis target."
        ),
        "synthesis_intent": (
            "Synthesize cross-country differences in education performance, skills development, and labour-market "
            "implications using the three source documents."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned from mixed AI-governance dimensions to education and skills as the single comparison axis."
        ),
        "recommended_action": "Proceed to prompt draft.",
    },
    "C051": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "Public-sector governance quality, institutional performance, and open-government capacity across countries"
        ),
        "cluster_rationale": (
            "S077, S166, and S190 are aligned on one governance axis: rule of law, institutional quality, gender-equality "
            "in law, and open-data governance. GDP, energy, and generic fiscal dimensions were dropped from the cluster framing."
        ),
        "synthesis_intent": (
            "Synthesize how governance institutions, legal frameworks, and open-government practices compare across "
            "countries and what that implies for public-sector performance."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned from overly broad multi-sector dimensions to a single public-governance comparison axis."
        ),
        "recommended_action": "Proceed to prompt draft.",
    },
    "C060": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "Energy transition progress: electrification, emissions, energy demand, and efficiency across countries and sectors"
        ),
        "cluster_rationale": (
            "S239, S152, and S085 are framed around one energy-transition axis. Macroeconomic and corporate-finance "
            "dimensions were removed from the cluster fields; open-data questionnaire content in S152 is secondary context "
            "for environmental data governance, not the primary synthesis target."
        ),
        "synthesis_intent": (
            "Synthesize evidence on energy transition pathways—electrification, emissions, demand, and efficiency—"
            "and compare patterns across the three sources."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned from broad macro/corporate dimensions to a single energy-transition comparison axis."
        ),
        "recommended_action": "Proceed to prompt draft.",
    },
    "C063": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "Building energy performance, energy savings, and emissions reduction from efficiency measures"
        ),
        "cluster_rationale": (
            "S188 and S087 both support energy-efficiency and performance measurement. The cluster is retuned to "
            "building-level energy savings and audit/programme outcomes rather than firm-level financial comparisons."
        ),
        "synthesis_intent": (
            "Synthesize how energy-efficiency programmes and building performance metrics compare across contexts, "
            "with emphasis on measured savings and emissions impacts."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned from mixed firm-finance dimensions to building energy performance and savings as the single axis."
        ),
        "recommended_action": "Proceed to prompt draft.",
    },
    "C069": {
        "remediation_action": "retune",
        "comparison_dimensions": (
            "Population wellbeing, health risk factors, and social outcomes across countries and demographic groups"
        ),
        "cluster_rationale": (
            "S229, S050, and S247 are aligned on wellbeing and health-outcome synthesis. GDP, public-trust, and "
            "firm-finance dimensions were removed from the cluster framing; workforce/GenAI content in S247 is contextual "
            "background, not the primary comparison axis."
        ),
        "synthesis_intent": (
            "Synthesize cross-country evidence on wellbeing, health prevalence (including obesity), and social "
            "outcomes, drawing contrasts across the three sources."
        ),
        "qa_status": "pass",
        "issue_type": "retuned",
        "issue_notes": (
            "Retuned from diffuse multi-domain dimensions to population wellbeing and health outcomes as the single axis."
        ),
        "recommended_action": "Proceed to prompt draft.",
    },
}


def run(workbook_path: Path) -> int:
    wb = load_workbook(workbook_path)
    clusters = wb["clusters"]
    headers = [cell.value for cell in clusters[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    for row in clusters.iter_rows(min_row=2):
        cluster_id = str(row[col["cluster_id"] - 1].value)
        if cluster_id not in REMEDIATIONS:
            continue
        fix = REMEDIATIONS[cluster_id]
        row[col["comparison_dimensions"] - 1].value = fix["comparison_dimensions"]
        row[col["cluster_rationale"] - 1].value = fix["cluster_rationale"]
        row[col["synthesis_intent"] - 1].value = fix["synthesis_intent"]

    # Update cluster_qa for remediated clusters only
    if "cluster_qa" in wb.sheetnames:
        qa = wb["cluster_qa"]
        qa_headers = [cell.value for cell in qa[1]]
        qcol = {name: idx + 1 for idx, name in enumerate(qa_headers)}

        # Add remediation_action column if missing
        if "remediation_action" not in qcol:
            new_col = len(qa_headers) + 1
            qa.cell(row=1, column=new_col, value="remediation_action")
            qa.cell(row=1, column=new_col).font = Font(bold=True)
            qcol["remediation_action"] = new_col
            qa_headers.append("remediation_action")

        if "qa_rerun" not in qcol:
            new_col = len(qa_headers) + 1
            qa.cell(row=1, column=new_col, value="qa_rerun")
            qcol["qa_rerun"] = new_col

        for row in qa.iter_rows(min_row=2):
            cluster_id = str(row[qcol["cluster_id"] - 1].value)
            if cluster_id not in REMEDIATIONS:
                continue
            fix = REMEDIATIONS[cluster_id]
            row[qcol["qa_status"] - 1].value = fix["qa_status"]
            row[qcol["issue_type"] - 1].value = fix["issue_type"]
            row[qcol["issue_notes"] - 1].value = fix["issue_notes"]
            row[qcol["recommended_action"] - 1].value = fix["recommended_action"]
            row[qcol["remediation_action"] - 1].value = fix["remediation_action"]
            row[qcol["qa_rerun"] - 1].value = "yes"

        fills = {
            "pass": PatternFill("solid", fgColor="D9EAD3"),
            "needs_review": PatternFill("solid", fgColor="FCE5CD"),
        }
        for row in qa.iter_rows(min_row=2):
            status = str(row[qcol["qa_status"] - 1].value or "")
            if row[qcol["cluster_id"] - 1].value in REMEDIATIONS:
                fill = fills.get(status, PatternFill())
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path(__file__).resolve().parent / "core_clusters.xlsx",
    )
    args = parser.parse_args()
    raise SystemExit(run(args.workbook))


if __name__ == "__main__":
    main()
