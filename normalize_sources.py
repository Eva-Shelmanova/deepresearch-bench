#!/usr/bin/env python3
"""Normalize Core and Set sources: inventory map -> one packet per source_id."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

BASE = Path(__file__).resolve().parent
STEP3 = BASE / "benchmark_step3_source_clustering.xlsx"
CORE_EXTRA = BASE / "core_extraction.xlsx"
RAW_CORE_DIR = BASE / "raw_text_core"
RAW_SET_DIR = BASE / "raw_text_set"
RAW_CORE_INDEX = BASE / "raw_sources_index.csv"
RAW_SET_INDEX = BASE / "raw_set_sources_index.csv"

OUT_CORE = BASE / "normalized_core"
OUT_SET = BASE / "normalized_set"
MAP_CSV = BASE / "source_to_file_map.csv"
CORE_DEDUP_CSV = BASE / "core_text_deduplication.csv"
REPORT_MD = BASE / "NORMALIZATION_REPORT.md"

FILE_HEADER_RE = re.compile(r"^===== FILE:\s*(.+?)\s*=====\s*$")
SPACE_RE = re.compile(r"[ \t\u00a0]+")
PAGE_MARKER_RE = re.compile(
    r"^(?:[-–—\s]*)?(?:page\s*)?\d{1,4}(?:\s*(?:of|/)\s*\d{1,4})?(?:[-–—\s]*)?$",
    re.I,
)
PDF_FOOTER_RE = re.compile(r"^[\W_]*\d{1,4}\s*[|]\s*.*$|^.*[|]\s*\d{1,4}[\W_]*$")
HYPHEN_BREAK_RE = re.compile(r"(\w)-\n(\w)")
SECTION_NUM_RE = re.compile(r"^(\d+(?:\.\d+)*)\.\s+(.+)$")
FRAGMENT_RE = re.compile(r"^===== FRAGMENT:", re.I)
SET_EXTRACTIONS = BASE / "set" / "extractions"
WIKI_STOP_HEADINGS = {
    "see also",
    "references",
    "explanatory notes",
    "notes",
    "further reading",
    "external links",
    "navigation menu",
    "categories",
    "hidden categories",
    "bibliography",
}
LICENSE_MARKERS = (
    "creative commons",
    "attribution 4.0",
    "all rights reserved",
    "unless otherwise noted",
)

WIKI_NAV_EXACT = {
    "jump to content",
    "main menu",
    "navigation",
    "search",
    "appearance",
    "tools",
    "personal tools",
    "create account",
    "log in",
    "move to sidebar",
    "hide",
    "show",
    "edit",
    "talk",
    "read",
    "view source",
    "view history",
    "main page",
    "contents",
    "current events",
    "random article",
    "about wikipedia",
    "contact us",
    "help",
    "learn to edit",
    "community portal",
    "recent changes",
    "upload file",
    "special pages",
    "donate",
    "contribute",
    "wikimedia commons",
    "wikimedia foundation",
    "mediawiki",
    "meta-wiki",
    "wikimedia outreach",
    "wikispecies",
    "wikidata",
    "wikifunctions",
    "wikivoyage",
    "wikinews",
    "wikiquote",
    "wikisource",
    "wikiversity",
    "wikibooks",
    "wiktionary",
    "print/export",
    "download as pdf",
    "printable version",
    "in other projects",
    "languages",
}


def load_step3_inventory() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(STEP3, read_only=True, data_only=True)
    core_rows: list[dict[str, Any]] = []
    set_rows: list[dict[str, Any]] = []

    ws = wb["Core Clusters"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        sid = str(row.get("Source ID") or "").strip()
        if sid:
            core_rows.append(
                {
                    "source_id": sid,
                    "family": "Core",
                    "title": str(row.get("Original text (EN)") or "").strip(),
                    "domain": row.get("Domain"),
                    "url": row.get("URL"),
                    "source_type": row.get("Source type"),
                    "source_subtype": row.get("Publication family"),
                    "status": row.get("Quality flag") or "in_step3_inventory",
                    "cluster_id": row.get("Cluster ID"),
                    "task_family": row.get("Task family"),
                    "notes": "",
                }
            )

    ws = wb["Set Clusters"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        sid = str(row.get("Source ID") or "").strip()
        if sid:
            set_rows.append(
                {
                    "source_id": sid,
                    "family": "Set",
                    "title": str(row.get("Original text (EN)") or "").strip(),
                    "domain": row.get("Domain"),
                    "url": row.get("URL"),
                    "source_type": row.get("Source type"),
                    "source_subtype": row.get("Source family"),
                    "status": row.get("QC status") or "in_step3_inventory",
                    "cluster_id": row.get("Cluster ID"),
                    "task_family": row.get("Task family"),
                    "notes": "",
                }
            )
    wb.close()
    return core_rows, set_rows


def load_core_extraction() -> dict[str, dict[str, Any]]:
    if not CORE_EXTRA.exists():
        return {}
    wb = openpyxl.load_workbook(CORE_EXTRA, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    out: dict[str, dict[str, Any]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        sid = str(row.get("source_id") or "").strip()
        if sid:
            out[sid] = row
    wb.close()
    return out


def load_raw_core_index() -> dict[str, list[dict[str, str]]]:
    by_source: dict[str, list[dict[str, str]]] = defaultdict(list)
    if not RAW_CORE_INDEX.exists():
        return by_source
    with RAW_CORE_INDEX.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            sid = (row.get("source_id") or "").strip()
            if sid:
                by_source[sid].append(row)
    return by_source


def load_raw_set_by_source() -> dict[str, list[Path]]:
    by_source: dict[str, list[Path]] = defaultdict(list)
    if RAW_SET_DIR.is_dir():
        for path in sorted(RAW_SET_DIR.glob("SET*.txt")):
            m = re.match(r"^(SET\d{3})_", path.name, re.I)
            if m:
                by_source[m.group(1).upper()].append(path)
    if RAW_SET_INDEX.exists():
        with RAW_SET_INDEX.open(encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                sid = (row.get("source_id") or "").strip()
                stem = (row.get("output_stem") or "").strip()
                if sid and stem:
                    p = RAW_SET_DIR / f"{stem}.txt"
                    if p.exists() and p not in by_source[sid]:
                        by_source[sid].append(p)
        for sid in by_source:
            by_source[sid] = sorted(set(by_source[sid]), key=lambda p: p.name)
    return by_source


def parse_raw_fragments(text: str) -> list[tuple[str, str]]:
    """Split combined raw text into (label, body) fragments."""
    parts: list[tuple[str, str]] = []
    current_label = "main"
    current_lines: list[str] = []

    for line in text.splitlines():
        m = FILE_HEADER_RE.match(line.strip())
        if m:
            if current_lines:
                parts.append((current_label, "\n".join(current_lines)))
            current_label = m.group(1).strip()
            current_lines = []
        else:
            current_lines.append(line)
    if current_lines:
        parts.append((current_label, "\n".join(current_lines)))
    return parts if parts else [("main", text)]


def read_concatenated_raw(paths: list[Path]) -> tuple[str, list[str]]:
    """Return combined raw text and fragment labels in order."""
    if not paths:
        return "", []
    chunks: list[str] = []
    labels: list[str] = []
    for path in paths:
        body = path.read_text(encoding="utf-8", errors="replace")
        frags = parse_raw_fragments(body)
        for label, frag in frags:
            labels.append(label if label != "main" else path.name)
            chunks.append(f"===== FRAGMENT: {label} =====\n{frag}")
    return "\n\n".join(chunks), labels


def is_page_noise(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if PAGE_MARKER_RE.fullmatch(s) and len(s) <= 20:
        return True
    if PDF_FOOTER_RE.fullmatch(s) and len(s) <= 180:
        return True
    if re.fullmatch(r"--\s*\d{1,4}\s+of\s+\d{1,4}\s*--", s, re.I):
        return True
    if re.fullmatch(r"\[\s*page\s+\d{1,4}\s*\]", s, re.I):
        return True
    return False


def is_wiki_nav(line: str) -> bool:
    s = line.strip()
    if not s or len(s) > 80:
        return False
    lower = s.lower()
    if lower in WIKI_NAV_EXACT:
        return True
    if lower.startswith("toggle ") and "contents" in lower:
        return True
    if re.fullmatch(r"(edit|talk|read|view|hide|show)", lower):
        return True
    return False


def repeated_noise_lines(lines: list[str]) -> set[str]:
    counts = Counter(line.strip() for line in lines if line.strip())
    noisy: set[str] = set()
    markers = ("©", "copyright", "all rights reserved", "isbn", "doi:", "shutterstock")
    for line, count in counts.items():
        lower = line.lower()
        if count >= 5 and len(line) <= 200 and any(m in lower for m in markers):
            noisy.add(line)
        elif count >= 10 and len(line) <= 120 and line.isupper():
            noisy.add(line)
    return noisy


def fix_hyphenation(text: str) -> str:
    return HYPHEN_BREAK_RE.sub(r"\1\2", text)


def normalize_line(line: str) -> str:
    line = line.replace("\ufeff", "").replace("\u00ad", "")
    return SPACE_RE.sub(" ", line).strip()


def block_to_markdown_table(lines: list[str]) -> str:
    rows: list[list[str]] = []
    for line in lines:
        if "\t" in line:
            cells = [c.strip() for c in line.split("\t")]
        elif "|" in line and line.count("|") >= 2:
            cells = [c.strip() for c in line.split("|")]
        else:
            cells = [c.strip() for c in re.split(r"\s{2,}", line) if c.strip()]
        if cells:
            rows.append(cells)
    if len(rows) < 2:
        return "\n".join(lines)
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    sep = ["---"] * width
    body = rows[1:]
    md_rows = [
        "| " + " | ".join(md_escape(c) for c in header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in body:
        md_rows.append("| " + " | ".join(md_escape(c) for c in row) + " |")
    return "\n".join(md_rows)


def md_escape(cell: str) -> str:
    return cell.replace("|", "\\|").replace("\n", " ")


def infer_heading_level(line: str) -> int | None:
    m = SECTION_NUM_RE.match(line)
    if m:
        depth = m.group(1).count(".") + 1
        return min(depth + 1, 4)
    letters = [c for c in line if c.isalpha()]
    if 3 <= len(line) <= 100 and letters:
        upper_ratio = sum(c.isupper() for c in letters) / len(letters)
        if upper_ratio > 0.8 and len(line.split()) <= 12:
            return 2
    if line.endswith(":") and len(line) <= 100:
        return 3
    return None


def structure_lines_to_markdown(lines: list[str], title: str, family: str) -> str:
    """Convert cleaned lines into markdown with headings and tables."""
    out: list[str] = [f"# {title or 'Untitled Source'}"]
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line.strip():
            i += 1
            continue
        if FRAGMENT_RE.match(line):
            out.append(f"\n## {line.replace('=', '').strip()}\n")
            i += 1
            continue

        # collect block
        block: list[str] = []
        while i < len(lines) and lines[i].strip():
            block.append(lines[i])
            i += 1

        if not block:
            continue

        if is_table_block(block):
            out.append("")
            out.append(block_to_markdown_table(block))
            out.append("")
            continue

        if len(block) == 1:
            level = infer_heading_level(block[0])
            if level:
                out.append("")
                out.append("#" * level + " " + block[0])
                out.append("")
            elif block[0].startswith(("- ", "* ", "• ")):
                out.append(block[0])
            else:
                out.append(block[0])
            continue

        # multi-line: heading + paragraph or list
        level = infer_heading_level(block[0])
        if level and len(block[0]) < 120:
            out.append("")
            out.append("#" * level + " " + block[0])
            out.append("")
            rest = " ".join(block[1:]) if len(block) > 1 else ""
            if rest:
                out.append(rest)
        elif all(b.startswith(("- ", "* ", "• ")) or len(b) < 80 for b in block):
            out.extend(block)
        else:
            out.append(" ".join(block))

    return "\n".join(out).strip()


def is_table_block(block: list[str]) -> bool:
    if len(block) < 2:
        return False
    tab_rows = sum(1 for b in block if "\t" in b)
    multi_space = sum(1 for b in block if len(re.findall(r"\s{2,}", b)) >= 2)
    pipe_rows = sum(1 for b in block if b.count("|") >= 2)
    return tab_rows >= 2 or multi_space >= max(2, len(block) // 2) or pipe_rows >= 2


def clean_raw_text(raw: str, family: str) -> tuple[str, list[str]]:
    """Return cleaned line list and normalization notes."""
    notes: list[str] = []
    text = fix_hyphenation(raw)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if HYPHEN_BREAK_RE.search(raw):
        notes.append("Merged hyphenation line-break artifacts.")

    raw_lines = [normalize_line(ln) for ln in text.splitlines()]
    repeated = repeated_noise_lines(raw_lines)
    if repeated:
        notes.append(f"Removed {len(repeated)} repeated footer/copyright/header lines.")

    lines: list[str] = []
    removed_nav = 0
    removed_page = 0
    prev_blank = False
    for line in raw_lines:
        if not line:
            if not prev_blank:
                lines.append("")
            prev_blank = True
            continue
        prev_blank = False
        if line in repeated:
            continue
        if is_page_noise(line):
            removed_page += 1
            continue
        if family == "Set" and is_wiki_nav(line):
            removed_nav += 1
            continue
        if family == "Set" and line.lower().startswith("retrieved from "):
            continue
        lines.append(line)

    if removed_page:
        notes.append(f"Removed {removed_page} page-number or footer markers.")
    if removed_nav:
        notes.append(f"Removed {removed_nav} navigation/UI lines (Set/HTML sources).")

    # collapse excess blanks
    collapsed: list[str] = []
    blank_run = 0
    for line in lines:
        if not line.strip():
            blank_run += 1
            if blank_run <= 1:
                collapsed.append("")
            continue
        blank_run = 0
        collapsed.append(line)

    if any(is_table_block([l]) for l in collapsed):
        notes.append("Preserved table structure where detected.")

    return collapsed, notes


def load_core_text_hash_groups() -> dict[str, list[str]]:
    """Group source_ids by identical extracted text hash from raw index."""
    by_hash: dict[str, list[str]] = defaultdict(list)
    if not RAW_CORE_INDEX.exists():
        return {}
    with RAW_CORE_INDEX.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            sid = (row.get("source_id") or "").strip()
            th = (row.get("text_sha256") or "").strip()
            if sid and th:
                by_hash[th].append(sid)
    return {h: sorted(set(ids)) for h, ids in by_hash.items() if len(set(ids)) > 1}


def resolve_core_canonical_id(
    source_id: str,
    core_extra: dict[str, dict[str, Any]],
    text_hash_groups: dict[str, list[str]],
    raw_index: dict[str, list[dict[str, str]]],
) -> tuple[str, str, bool]:
    """
    Return (canonical_source_id, duplicate_group, is_text_duplicate).
    Prefer core_extraction canonical; fall back to lowest S### in same text-hash group.
    """
    extra = core_extra.get(source_id, {})
    canonical = str(extra.get("canonical_source_id") or source_id).strip()
    dup_group = "none"
    is_dup = canonical != source_id

    if extra.get("duplicate_file_group"):
        dup_group = str(extra.get("duplicate_file_group"))
    elif is_dup:
        dup_group = f"variant_of:{canonical}"
    elif extra.get("duplicate_reason"):
        dup_group = str(extra.get("duplicate_reason"))

    # Cross-check text hash for step3 ids not marked as variants in core_extraction
    records = raw_index.get(source_id, [])
    text_hash = ""
    if records:
        text_hash = (records[0].get("text_sha256") or "").strip()

    if text_hash and text_hash in text_hash_groups:
        group = text_hash_groups[text_hash]
        if len(group) > 1:
            hash_canonical = sorted(group)[0]
            if source_id != hash_canonical:
                is_dup = True
                canonical = hash_canonical
                if dup_group == "none":
                    dup_group = f"text_hash_group:{text_hash[:16]}"
            elif source_id == hash_canonical:
                dup_group = (
                    dup_group
                    if dup_group != "none"
                    else f"canonical_for:{','.join(g for g in group if g != source_id)}"
                )

    return canonical, dup_group, is_dup


def build_variant_alias_body(meta: dict[str, Any], canonical_id: str) -> str:
    ref = f"normalized_core/{canonical_id}.normalized.md"
    lines = [
        "This inventory row is a **duplicate variant** of a document whose text is stored once.",
        "",
        f"**Canonical source ID:** `{canonical_id}`",
        f"**Canonical normalized packet:** `{ref}`",
        "",
        "The underlying extracted text is identical to the canonical source. "
        "Do not treat this row as independent evidence in clustering or model tasks.",
        "",
        "### Variant-specific inventory metadata (differs from canonical)",
        "",
        f"- **Title (step3):** {meta.get('title') or '(none)'}",
        f"- **URL (step3):** {meta.get('url') or '(none)'}",
        f"- **Source type (step3):** {meta.get('source_type') or '(none)'}",
        f"- **Source subtype (step3):** {meta.get('source_subtype') or '(none)'}",
        f"- **Cluster ID (step3):** {meta.get('cluster_id') or '(none)'}",
    ]
    return "\n".join(lines)


def build_metadata_header(meta: dict[str, Any]) -> str:
    lines = ["# Normalized Source Packet", ""]
    for key in (
        "source_id",
        "canonical_source_id",
        "family",
        "title",
        "domain",
        "url",
        "source_type",
        "source_subtype",
        "status",
        "duplicate_group",
        "is_text_duplicate",
        "normalized_text_ref",
    ):
        val = meta.get(key)
        if val is not None and val != "":
            lines.append(f"- **{key}:** {val}")
    if meta.get("packet_role"):
        lines.append(f"- **packet_role:** {meta['packet_role']}")
    if meta.get("raw_fragment_files"):
        lines.append(f"- **raw_fragment_count:** {len(meta['raw_fragment_files'])}")
    return "\n".join(lines) + "\n\n---\n\n"


def infer_title(inv: dict[str, Any], raw_text: str, family: str) -> str:
    title = str(inv.get("title") or "").strip()
    if " — " in title:
        title = title.split(" — ", 1)[0].strip()
    if title and len(title) < 200 and not title.endswith("."):
        return title
    if family == "Core":
        for line in raw_text.splitlines()[:80]:
            line = normalize_line(line)
            if "Survey on Drivers" in line or "OECD" in line and len(line) < 120:
                return line
            if len(line) > 20 and len(line) < 120 and "©" not in line:
                if line.isupper() or line[0].isupper():
                    return line
    if family == "Set":
        for line in raw_text.splitlines()[:40]:
            line = normalize_line(line)
            if " - Wikipedia" in line:
                return line.replace(" - Wikipedia", "").strip()
    return title or str(inv.get("source_id"))


def load_set_extraction_payloads(set_id: str) -> list[dict[str, Any]]:
    folder = SET_EXTRACTIONS / set_id
    if not folder.is_dir():
        return []
    payloads: list[dict[str, Any]] = []
    for path in sorted(folder.glob("*.json")):
        try:
            payloads.append(json.loads(path.read_text(encoding="utf-8")))
        except json.JSONDecodeError:
            continue
    return payloads


def entities_to_markdown(entities: list[dict[str, Any]]) -> str:
    if not entities:
        return ""
    table_rows = [e for e in entities if e.get("cells")]
    link_rows = [e for e in entities if not e.get("cells") and e.get("url")]

    parts: list[str] = []
    if table_rows:
        current_caption = None
        blocks: list[list[str]] = []
        for ent in table_rows:
            caption = ent.get("caption") or ""
            cells = ent.get("cells") or []
            if not cells:
                continue
            if caption and caption != current_caption:
                if blocks:
                    parts.append(cells_to_markdown_table(blocks))
                    blocks = []
                parts.append(f"### {caption}")
                current_caption = caption
            blocks.append(cells)
        if blocks:
            parts.append(cells_to_markdown_table(blocks))

    if link_rows:
        parts.append("### Linked items")
        for ent in link_rows[:500]:
            label = md_escape(str(ent.get("label") or ""))
            url = str(ent.get("url") or "").strip()
            if label and url:
                parts.append(f"- {label} ({url})")
            elif label:
                parts.append(f"- {label}")
        if len(link_rows) > 500:
            parts.append(f"- … and {len(link_rows) - 500} more linked items")

    return "\n\n".join(parts).strip()


def cells_to_markdown_table(rows: list[list[str]]) -> str:
    parsed = [[md_escape(str(c)) for c in row] for row in rows if row]
    if len(parsed) < 2:
        return "\n".join(" | ".join(r) for r in parsed)
    width = max(len(r) for r in parsed)
    parsed = [r + [""] * (width - len(r)) for r in parsed]
    header = parsed[0]
    sep = ["---"] * width
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in parsed[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def trim_wiki_tail(text: str) -> tuple[str, bool]:
    lines = text.splitlines()
    out: list[str] = []
    stopped = False
    for line in lines:
        norm = normalize_line(line).lower()
        if norm in WIKI_STOP_HEADINGS or norm.startswith("category:"):
            stopped = True
            break
        out.append(line)
    return "\n".join(out), stopped


def strip_license_block(lines: list[str]) -> tuple[list[str], bool]:
    """Remove leading license/copyright block before Foreword or chapter 1."""
    joined = "\n".join(lines)
    lower = joined.lower()
    if not any(m in lower[:4000] for m in LICENSE_MARKERS):
        return lines, False
    for i, line in enumerate(lines):
        norm = normalize_line(line).lower()
        if norm in {"foreword", "executive summary", "introduction", "abstract"}:
            return lines[i:], True
        if SECTION_NUM_RE.match(normalize_line(line)) or re.match(r"^chapter\s+\d", norm):
            return lines[i:], True
    return lines, False


def resolve_core_raw_path(
    source_id: str,
    canonical_id: str,
) -> list[Path]:
    """Always read canonical raw text for normalization (text-level dedup)."""
    paths: list[Path] = []
    for folder in (RAW_CORE_DIR, BASE / "raw_text"):
        primary = folder / f"{canonical_id}.txt"
        if primary.exists():
            paths.append(primary)
            break
    return paths


def normalize_set_with_extractions(
    inv: dict[str, Any],
    raw_paths: list[Path],
    payloads: list[dict[str, Any]],
) -> tuple[str, list[str]]:
    notes: list[str] = []
    raw_text, fragment_labels = read_concatenated_raw(raw_paths)
    title = infer_title(inv, raw_text, "Set")

    if len(raw_paths) > 1 or len(fragment_labels) > 1:
        notes.append(
            f"Combined {len(raw_paths)} raw file(s) / {len(fragment_labels)} fragment(s) into one source packet."
        )
    if payloads:
        notes.append(
            f"Preserved structured rows from {len(payloads)} extraction JSON file(s) as Markdown tables/lists."
        )

    sections: list[str] = [f"# {title}"]

    # Lead from first payload or cleaned raw intro
    leads: list[str] = []
    for payload in payloads[:3]:
        for para in payload.get("lead_paragraphs") or []:
            cleaned_para, _ = clean_raw_text(para, "Set")
            para = " ".join(ln for ln in cleaned_para if ln)
            if para and para not in leads:
                leads.append(para)
    if leads:
        sections.extend(["", "## Lead", "", "\n\n".join(leads)])

    if payloads:
        for payload in payloads:
            page_title = payload.get("page_title") or payload.get("source_file")
            entities = payload.get("entities") or []
            md = entities_to_markdown(entities)
            if md:
                sections.extend(["", f"## {page_title}", "", md])
    elif raw_text.strip():
        cleaned, clean_notes = clean_raw_text(raw_text, "Set")
        notes.extend(clean_notes)
        cleaned, _ = trim_wiki_tail("\n".join(cleaned))
        body = structure_lines_to_markdown(
            cleaned.splitlines() if isinstance(cleaned, str) else cleaned,
            title,
            "Set",
        )
        sections.extend(["", body])

    return "\n\n".join(sections).strip(), notes


def normalize_source(
    inv: dict[str, Any],
    raw_paths: list[Path],
    core_extra: dict[str, dict[str, Any]] | None = None,
    set_payloads: list[dict[str, Any]] | None = None,
) -> tuple[str, dict[str, Any], list[str]]:
    notes: list[str] = []
    family = inv["family"]
    source_id = inv["source_id"]

    canonical = str(inv.get("canonical_source_id") or source_id)
    duplicate_group = str(inv.get("duplicate_group") or "none")
    is_text_duplicate = bool(inv.get("is_text_duplicate"))

    if family == "Core" and core_extra:
        extra = core_extra.get(source_id, {})
        if extra.get("final_decision"):
            inv["status"] = str(extra.get("final_decision"))
        if is_text_duplicate:
            notes.append(
                f"Text-level duplicate of {canonical}; body not duplicated (see canonical packet)."
            )

    raw_text, fragment_labels = read_concatenated_raw(raw_paths)

    if family == "Set":
        body_md, set_notes = normalize_set_with_extractions(
            inv, raw_paths, set_payloads or []
        )
        notes.extend(set_notes)
        title = infer_title(inv, raw_text, family)
    elif not raw_text.strip():
        notes.append("No raw text available for this source_id.")
        body_md = "_No raw text was available at normalization time._"
        title = infer_title(inv, "", family)
    else:
        if len(raw_paths) > 1 or len(fragment_labels) > 1:
            notes.append(
                f"Combined {len(raw_paths)} raw file(s) / {len(fragment_labels)} fragment(s) into one source packet."
            )
        cleaned_lines, clean_notes = clean_raw_text(raw_text, family)
        notes.extend(clean_notes)
        trimmed, stopped = trim_wiki_tail("\n".join(cleaned_lines))
        if stopped:
            notes.append(
                "Trimmed trailing Wikipedia sections (See also, References, categories)."
            )
        cleaned_lines = trimmed.splitlines()
        stripped, removed_license = strip_license_block(cleaned_lines)
        if removed_license:
            notes.append("Removed leading copyright/license boilerplate block.")
            cleaned_lines = stripped
        title = infer_title(inv, raw_text, family)
        body_md = structure_lines_to_markdown(cleaned_lines, title, family)
        notes.append("Applied conservative cleanup without summarization or reordering.")

    meta: dict[str, Any] = {
        "source_id": source_id,
        "canonical_source_id": canonical,
        "family": family,
        "title": title,
        "domain": inv.get("domain"),
        "url": inv.get("url"),
        "source_type": inv.get("source_type"),
        "source_subtype": inv.get("source_subtype"),
        "status": inv.get("status"),
        "duplicate_group": duplicate_group,
        "is_text_duplicate": is_text_duplicate,
        "normalized_text_ref": (
            f"normalized_core/{canonical}.normalized.md"
            if is_text_duplicate and family == "Core"
            else ""
        ),
        "packet_role": (
            "alias"
            if is_text_duplicate and family == "Core"
            else ("canonical" if family == "Core" else "primary")
        ),
        "cluster_id": inv.get("cluster_id"),
        "task_family": inv.get("task_family"),
        "raw_fragment_files": [p.name for p in raw_paths],
        "fragment_labels": fragment_labels,
        "normalization_notes": notes,
        "normalized_at_utc": datetime.now(timezone.utc).isoformat(),
        "raw_char_count": len(raw_text),
        "normalized_char_count": len(body_md),
    }

    header = build_metadata_header(meta)
    full_md = header + body_md
    return full_md, meta, notes


def write_packet(
    out_dir: Path,
    source_id: str,
    md_content: str,
    meta: dict[str, Any],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    md_path = out_dir / f"{source_id}.normalized.md"
    json_path = out_dir / f"{source_id}.metadata.json"
    md_path.write_text(md_content, encoding="utf-8")
    body = md_content.split("---\n\n", 1)[-1].strip() if "---\n\n" in md_content else md_content
    payload = {
        **meta,
        "normalized_text_path": str(md_path.relative_to(BASE)),
        "normalized_text_preview": body[:2000] + ("…" if len(body) > 2000 else ""),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def validate_core_text_dedup(out_dir: Path) -> dict[str, Any]:
    """Verify canonical packets hold unique text; variants are aliases only."""
    import hashlib

    def body_hash(path: Path) -> str:
        text = path.read_text(encoding="utf-8")
        body = text.split("---\n\n", 1)[-1].strip() if "---\n\n" in text else text
        return hashlib.sha256(body.encode()).hexdigest()

    canonical_hashes: dict[str, str] = {}
    variant_with_full_dup: list[str] = []
    alias_ids: list[str] = []

    for path in sorted(out_dir.glob("*.normalized.md")):
        sid = path.stem.replace(".normalized", "")
        meta_path = out_dir / f"{sid}.metadata.json"
        role = "canonical"
        is_dup = False
        cid = sid
        if meta_path.exists():
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
            role = meta.get("packet_role", role)
            is_dup = bool(meta.get("is_text_duplicate"))
            cid = meta.get("canonical_source_id", sid)

        h = body_hash(path)
        if role == "alias" or is_dup:
            alias_ids.append(sid)
            canon_path = out_dir / f"{cid}.normalized.md"
            if canon_path.exists() and h == body_hash(canon_path):
                variant_with_full_dup.append(sid)
        else:
            if cid in canonical_hashes and canonical_hashes[cid] != h:
                pass  # same canonical id shouldn't appear twice as canonical role
            canonical_hashes[cid] = h

    unique_bodies = len(set(canonical_hashes.values()))
    return {
        "unique_canonical_bodies": unique_bodies,
        "canonical_packets": len(canonical_hashes),
        "alias_packets": len(alias_ids),
        "variants_still_duplicating_body": variant_with_full_dup,
        "pass": len(variant_with_full_dup) == 0,
    }


def main() -> int:
    core_inv, set_inv = load_step3_inventory()
    core_extra = load_core_extraction()
    raw_core_index = load_raw_core_index()
    raw_set_by_source = load_raw_set_by_source()
    text_hash_groups = {
        h: ids for h, ids in load_core_text_hash_groups().items()
    }

    map_rows: list[dict[str, Any]] = []
    dedup_rows: list[dict[str, Any]] = []
    stats = Counter()

    # Cache: canonical_id -> (full_md, meta) for canonical body text
    core_body_cache: dict[str, tuple[str, dict[str, Any]]] = {}

    # Resolve canonical mapping for every Core inventory row
    for inv in core_inv:
        sid = inv["source_id"]
        canonical, dup_group, is_dup = resolve_core_canonical_id(
            sid, core_extra, text_hash_groups, raw_core_index
        )
        inv["canonical_source_id"] = canonical
        inv["duplicate_group"] = dup_group
        inv["is_text_duplicate"] = is_dup

    # Process canonical sources first, then variants
    def core_order(inv: dict[str, Any]) -> tuple:
        sid = inv["source_id"]
        is_var = 1 if inv.get("is_text_duplicate") else 0
        return (is_var, inv["canonical_source_id"], sid)

    for inv in sorted(core_inv, key=core_order):
        sid = inv["source_id"]
        canonical = inv["canonical_source_id"]
        is_dup = inv["is_text_duplicate"]
        paths = resolve_core_raw_path(sid, canonical)

        dedup_rows.append(
            {
                "source_id": sid,
                "canonical_source_id": canonical,
                "is_text_duplicate": is_dup,
                "duplicate_group": inv["duplicate_group"],
                "raw_text_file_used": paths[0].name if paths else "",
                "packet_role": "alias" if is_dup else "canonical",
            }
        )

        map_rows.append(
            {
                "source_id": sid,
                "family": "Core",
                "canonical_source_id": canonical,
                "title": inv.get("title"),
                "domain": inv.get("domain"),
                "url": inv.get("url"),
                "source_type": inv.get("source_type"),
                "source_subtype": inv.get("source_subtype"),
                "status": inv.get("status"),
                "duplicate_group": inv["duplicate_group"],
                "is_text_duplicate": is_dup,
                "raw_files": paths[0].name if paths else "",
                "raw_file_count": len(paths),
                "mapped": "yes" if paths else "missing",
                "packet_role": "alias" if is_dup else "canonical",
                "normalized_text_ref": (
                    f"normalized_core/{canonical}.normalized.md" if is_dup else ""
                ),
                "normalized_output": f"normalized_core/{sid}.normalized.md",
            }
        )

        if not paths:
            stats["core_missing_raw"] += 1
            md, meta, _ = normalize_source(inv, [], core_extra)
            meta["status"] = "missing_raw_text"
            write_packet(OUT_CORE, sid, md, meta)
            continue

        if is_dup:
            if canonical not in core_body_cache:
                stats["core_variant_before_canonical"] += 1
            canon_md, canon_meta = core_body_cache.get(canonical, ("", {}))
            if not canon_md:
                canon_paths = resolve_core_raw_path(canonical, canonical)
                canon_inv = next(
                    (r for r in core_inv if r["source_id"] == canonical),
                    {**inv, "source_id": canonical, "is_text_duplicate": False},
                )
                canon_inv["canonical_source_id"] = canonical
                canon_inv["duplicate_group"] = "none"
                canon_inv["is_text_duplicate"] = False
                canon_md, canon_meta, _ = normalize_source(canon_inv, canon_paths, core_extra)
                core_body_cache[canonical] = (canon_md, canon_meta)

            meta = dict(canon_meta)
            meta.update(
                {
                    "source_id": sid,
                    "canonical_source_id": canonical,
                    "is_text_duplicate": True,
                    "duplicate_group": inv["duplicate_group"],
                    "packet_role": "alias",
                    "normalized_text_ref": f"normalized_core/{canonical}.normalized.md",
                    "title": (str(inv.get("title") or "").split(" — ")[0].strip()
                              or canon_meta.get("title")),
                    "url": inv.get("url"),
                    "source_type": inv.get("source_type"),
                    "source_subtype": inv.get("source_subtype"),
                    "cluster_id": inv.get("cluster_id"),
                    "normalization_notes": [
                        f"Alias packet only; full text stored under {canonical}.",
                        *canon_meta.get("normalization_notes", [])[:3],
                    ],
                }
            )
            header = build_metadata_header(meta)
            header = header.replace(
                "# Normalized Source Packet",
                "# Normalized Source Packet (alias — text duplicate)",
                1,
            )
            body = build_variant_alias_body(meta, canonical)
            md = header + body
            stats["core_alias"] += 1
            write_packet(OUT_CORE, sid, md, meta)
            continue

        # Canonical packet: normalize once per canonical_id
        if canonical in core_body_cache:
            md, meta = core_body_cache[canonical]
            stats["core_canonical_reuse"] += 1
        else:
            canon_inv = dict(inv)
            canon_inv["is_text_duplicate"] = False
            md, meta, _ = normalize_source(canon_inv, paths, core_extra)
            meta["packet_role"] = "canonical"
            meta["is_text_duplicate"] = False
            meta["normalized_text_ref"] = ""
            core_body_cache[canonical] = (md, meta)
            stats["core_canonical_new"] += 1

        write_packet(OUT_CORE, sid, md, meta)
        stats["core_ok"] += 1

    for inv in set_inv:
        sid = inv["source_id"]
        paths = raw_set_by_source.get(sid, [])
        inv["canonical_source_id"] = sid
        inv["duplicate_group"] = "none"

        payloads = load_set_extraction_payloads(sid)
        map_rows.append(
            {
                "source_id": sid,
                "family": "Set",
                "canonical_source_id": sid,
                "title": inv.get("title"),
                "domain": inv.get("domain"),
                "url": inv.get("url"),
                "source_type": inv.get("source_type"),
                "source_subtype": inv.get("source_subtype"),
                "status": inv.get("status"),
                "duplicate_group": "none",
                "raw_files": "; ".join(p.name for p in paths[:5])
                + ("; ..." if len(paths) > 5 else ""),
                "raw_file_count": len(paths),
                "extraction_json_count": len(payloads),
                "mapped": "yes" if paths else "missing",
                "normalized_output": f"normalized_set/{sid}.normalized.md",
            }
        )

        if not paths:
            stats["set_missing_raw"] += 1
            md, meta, _ = normalize_source(inv, [], None, payloads)
            meta["status"] = "missing_raw_text"
        else:
            md, meta, _ = normalize_source(inv, paths, None, payloads)
            stats["set_ok"] += 1

        write_packet(OUT_SET, sid, md, meta)

    # Write map CSV
    with MAP_CSV.open("w", encoding="utf-8", newline="") as f:
        fieldnames = [
            "source_id",
            "family",
            "canonical_source_id",
            "title",
            "domain",
            "url",
            "source_type",
            "source_subtype",
            "status",
            "duplicate_group",
            "raw_file_count",
            "extraction_json_count",
            "raw_files",
            "mapped",
            "normalized_output",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(map_rows)

    with CORE_DEDUP_CSV.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "source_id",
                "canonical_source_id",
                "is_text_duplicate",
                "duplicate_group",
                "raw_text_file_used",
                "packet_role",
            ],
        )
        writer.writeheader()
        writer.writerows(dedup_rows)

    # Validation
    dedup_check = validate_core_text_dedup(OUT_CORE)
    core_files = list(OUT_CORE.glob("*.normalized.md"))
    set_files = list(OUT_SET.glob("*.normalized.md"))
    empty_core = [p.stem.replace(".normalized", "") for p in core_files if p.stat().st_size < 200]
    empty_set = [p.stem.replace(".normalized", "") for p in set_files if p.stat().st_size < 200]

    report = [
        "# Normalization Report",
        "",
        f"Generated: {datetime.now(timezone.utc).isoformat()}",
        "",
        "## Counts",
        "",
        f"- Step3 Core sources: **{len(core_inv)}**",
        f"- Step3 Set sources: **{len(set_inv)}**",
        f"- `normalized_core/*.normalized.md`: **{len(core_files)}**",
        f"- `normalized_set/*.normalized.md`: **{len(set_files)}**",
        f"- Core with raw text: **{stats['core_ok']}**",
        f"- Core missing raw: **{stats['core_missing_raw']}**",
        f"- Core canonical packets (unique text): **{stats['core_canonical_new']}**",
        f"- Core alias packets (no duplicated body): **{stats['core_alias']}**",
        f"- Set with raw text: **{stats['set_ok']}**",
        f"- Set missing raw: **{stats['set_missing_raw']}**",
        "",
        "## Core text-level deduplication",
        "",
        f"- Unique canonical body texts on disk: **{dedup_check['unique_canonical_bodies']}**",
        f"- Alias packets (metadata only): **{dedup_check['alias_packets']}**",
        f"- Variants still copying full canonical body: **{len(dedup_check['variants_still_duplicating_body'])}** "
        f"{'PASS' if dedup_check['pass'] else 'FAIL'}",
        "",
        "See `core_text_deduplication.csv` for per-source canonical mapping.",
        "",
        "## Validation",
        "",
        f"- Core inventory vs normalized files: {'PASS' if len(core_files) == len(core_inv) else 'FAIL'} ({len(core_files)} vs {len(core_inv)})",
        f"- Core text deduplication: {'PASS' if dedup_check['pass'] else 'FAIL'}",
        f"- Set inventory vs normalized files: {'PASS' if len(set_files) == len(set_inv) else 'FAIL'} ({len(set_files)} vs {len(set_inv)})",
        "",
        "## Missing raw text (Core)",
        "",
    ]
    missing_core = [r["source_id"] for r in map_rows if r["family"] == "Core" and r["mapped"] == "missing"]
    report.extend(f"- {sid}" for sid in missing_core or ["(none)"])
    report.extend(["", "## Missing raw text (Set)", ""])
    missing_set = [r["source_id"] for r in map_rows if r["family"] == "Set" and r["mapped"] == "missing"]
    report.extend(f"- {sid}" for sid in missing_set or ["(none)"])
    report.extend(
        [
            "",
            "## Very small outputs (<200 bytes)",
            "",
            f"- Core: {', '.join(empty_core) if empty_core else '(none)'}",
            f"- Set: {', '.join(empty_set) if empty_set else '(none)'}",
        ]
    )

    REPORT_MD.write_text("\n".join(report) + "\n", encoding="utf-8")

    print(f"Wrote {len(core_files)} Core + {len(set_files)} Set normalized packets")
    print(f"Map: {MAP_CSV}")
    print(f"Report: {REPORT_MD}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
