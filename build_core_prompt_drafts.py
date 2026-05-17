#!/usr/bin/env python3
"""Generate Core synthesis prompt drafts from QA-ready clusters."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font


EXCERPT_PREVIEW_WORDS = 180


@dataclass
class ClusterRow:
    cluster_id: str
    source_ids: str
    canonical_source_ids: str
    domain: str
    expected_output_type: str
    comparison_dimensions: str
    synthesis_intent: str
    cluster_rationale: str


def clean_excel(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    stripped = cleaned.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        cleaned = cleaned[: len(cleaned) - len(stripped)] + "'" + stripped
    return cleaned


def preview_excerpt(text: str, max_words: int = EXCERPT_PREVIEW_WORDS) -> str:
    words = text.split()
    if len(words) <= max_words:
        return text.strip()
    return " ".join(words[:max_words]).strip() + " [...]"


def main_comparison_axis(dimensions: str) -> str:
    """Use the first clause as the primary axis when dimensions were retuned to one lead phrase."""
    parts = [part.strip() for part in dimensions.split(";") if part.strip()]
    return parts[0] if parts else dimensions.strip()


def build_prompt(cluster: ClusterRow) -> str:
    sources = [sid.strip() for sid in cluster.source_ids.split(";") if sid.strip()]
    source_list = ", ".join(sources)
    axis = main_comparison_axis(cluster.comparison_dimensions)

    lines = [
        "CORE SYNTHESIS TASK",
        "",
        f"Domain: {cluster.domain}",
        f"Sources (use only these): {source_list}",
        f"Expected output type: {cluster.expected_output_type}",
        "",
        "Main comparison axis:",
        axis,
        "",
        "Task:",
        (
            "Write a synthesis of 120–180 words that integrates evidence from all listed sources. "
            "Do not write one paragraph per source; combine insights into a single comparative argument."
        ),
        "",
        "Your synthesis must:",
        f"- Stay focused on the main comparison axis: {axis}.",
        "- Cite at least two quantitative facts or metrics that appear in the source material.",
        "- Identify at least one clear difference or contrast between the sources.",
        "- Include at least one sentence on implications or conclusions supported by the evidence.",
        "- Use only information grounded in the provided excerpts; do not invent statistics, names, or citations.",
        "",
        "Synthesis intent (for reviewers):",
        cluster.synthesis_intent,
    ]
    return "\n".join(lines)


def load_clusters(path: Path) -> list[ClusterRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["clusters"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}

    clusters: list[ClusterRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["cluster_status"]] or "") != "ready":
            continue
        clusters.append(
            ClusterRow(
                cluster_id=str(row[idx["cluster_id"]]),
                source_ids=str(row[idx["source_ids"]]),
                canonical_source_ids=str(row[idx["canonical_source_ids"]]),
                domain=str(row[idx["domain"]]),
                expected_output_type=str(row[idx["expected_output_type"]]),
                comparison_dimensions=str(row[idx["comparison_dimensions"]]),
                synthesis_intent=str(row[idx["synthesis_intent"]]),
                cluster_rationale=str(row[idx["cluster_rationale"]]),
            )
        )
    wb.close()
    return clusters


def load_excerpts(path: Path) -> dict[str, tuple[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["core_extraction"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}

    excerpts: dict[str, tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[idx["final_decision"]] or "") != "COMPLETE":
            continue
        source_id = str(row[idx["source_id"]])
        excerpts[source_id] = (
            str(row[idx["file_names"]] or ""),
            str(row[idx["original_excerpt"]] or ""),
        )
    wb.close()
    return excerpts


def write_workbook(
    clusters: list[ClusterRow],
    excerpts: dict[str, tuple[str, str]],
    output_path: Path,
) -> None:
    wb = Workbook()
    drafts = wb.active
    drafts.title = "prompt_drafts"

    draft_headers = [
        "cluster_id",
        "source_ids",
        "canonical_source_ids",
        "domain",
        "expected_output_type",
        "main_comparison_axis",
        "comparison_dimensions",
        "target_word_count_min",
        "target_word_count_max",
        "prompt_draft",
        "prompt_status",
    ]
    drafts.append(draft_headers)

    material = wb.create_sheet("source_material")
    material.append(["cluster_id", "source_id", "file_names", "excerpt_preview"])

    for cluster in clusters:
        axis = main_comparison_axis(cluster.comparison_dimensions)
        prompt = build_prompt(cluster)
        drafts.append(
            [
                cluster.cluster_id,
                cluster.source_ids,
                cluster.canonical_source_ids,
                cluster.domain,
                cluster.expected_output_type,
                axis,
                cluster.comparison_dimensions,
                120,
                180,
                prompt,
                "draft",
            ]
        )

        for source_id in [s.strip() for s in cluster.source_ids.split(";") if s.strip()]:
            file_names, excerpt = excerpts.get(source_id, ("", ""))
            material.append(
                [
                    cluster.cluster_id,
                    source_id,
                    file_names,
                    preview_excerpt(excerpt),
                ]
            )

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    summary_rows = [
        ("clusters_with_prompt_drafts", len(clusters)),
        ("target_output_words", "120-180"),
        ("prompt_status", "draft (no rubrics yet)"),
        ("source_material_rows", material.max_row - 1),
    ]
    for metric, value in summary_rows:
        summary.append([metric, value])

    for ws in (drafts, material, summary):
        for row in ws.iter_rows():
            for cell in row:
                cell.value = clean_excel(cell.value)
                if isinstance(cell.value, str):
                    cell.data_type = "s"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]:
            cell.font = Font(bold=True)

    drafts.column_dimensions["A"].width = 12
    drafts.column_dimensions["B"].width = 22
    drafts.column_dimensions["C"].width = 22
    drafts.column_dimensions["D"].width = 28
    drafts.column_dimensions["E"].width = 26
    drafts.column_dimensions["F"].width = 42
    drafts.column_dimensions["G"].width = 50
    drafts.column_dimensions["J"].width = 90
    material.column_dimensions["D"].width = 100
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 24

    drafts.freeze_panes = "A2"
    material.freeze_panes = "A2"
    wb.save(output_path)


def run(clusters_path: Path, extraction_path: Path, output_path: Path) -> int:
    clusters = load_clusters(clusters_path)
    excerpts = load_excerpts(extraction_path)
    write_workbook(clusters, excerpts, output_path)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    base = Path(__file__).resolve().parent
    parser.add_argument("--clusters", type=Path, default=base / "core_clusters.xlsx")
    parser.add_argument("--extraction", type=Path, default=base / "core_extraction.xlsx")
    parser.add_argument("--output", type=Path, default=base / "core_prompt_drafts.xlsx")
    args = parser.parse_args()
    raise SystemExit(run(args.clusters, args.extraction, args.output))


if __name__ == "__main__":
    main()
