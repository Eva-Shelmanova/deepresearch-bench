#!/usr/bin/env python3
"""Build Core source clusters from COMPLETE rows in core_extraction.xlsx."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill


WORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9_-]+")


@dataclass(frozen=True)
class SourceRow:
    source_id: str
    canonical_source_id: str
    domain: str
    source_type: str
    comparison_dimensions: str
    synthesis_potential: str
    likely_report_theme: str
    word_count: int


@dataclass
class Cluster:
    cluster_id: str
    sources: list[SourceRow]
    cluster_status: str
    cluster_rationale: str
    expected_output_type: str
    comparison_dimensions: str
    synthesis_intent: str


STOPWORDS = {
    "and",
    "the",
    "for",
    "with",
    "using",
    "source",
    "evidence",
    "comparative",
    "comparison",
    "report",
    "domain",
    "indicators",
    "reported",
    "trends",
    "quantitative",
}


def clean_excel_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    stripped = cleaned.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        cleaned = cleaned[: len(cleaned) - len(stripped)] + "'" + stripped
    return cleaned


def tokens(text: str) -> set[str]:
    return {
        token.lower()
        for token in WORD_RE.findall(text or "")
        if len(token) > 2 and token.lower() not in STOPWORDS
    }


def load_complete_sources(path: Path) -> tuple[list[SourceRow], Counter[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["core_extraction"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}

    complete: list[SourceRow] = []
    excluded_status_counts: Counter[str] = Counter()
    seen_canonical: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        final_decision = str(row[idx["final_decision"]] or "")
        if final_decision != "COMPLETE":
            excluded_status_counts[final_decision] += 1
            continue

        canonical = str(row[idx["canonical_source_id"]] or row[idx["source_id"]])
        # Defensive duplicate guard: only one COMPLETE row per canonical source may enter clustering.
        if canonical in seen_canonical:
            excluded_status_counts["duplicate_complete_canonical_guard"] += 1
            continue
        seen_canonical.add(canonical)

        complete.append(
            SourceRow(
                source_id=str(row[idx["source_id"]]),
                canonical_source_id=canonical,
                domain=str(row[idx["domain"]] or ""),
                source_type=str(row[idx["source_type"]] or ""),
                comparison_dimensions=str(row[idx["comparison_dimensions"]] or ""),
                synthesis_potential=str(row[idx["synthesis_potential"]] or ""),
                likely_report_theme=str(row[idx["likely_report_theme"]] or ""),
                word_count=int(row[idx["word_count"]] or 0),
            )
        )

    wb.close()
    return complete, excluded_status_counts


def complementarity_score(a: SourceRow, b: SourceRow) -> float:
    dim_a = tokens(a.comparison_dimensions)
    dim_b = tokens(b.comparison_dimensions)
    theme_a = tokens(a.likely_report_theme)
    theme_b = tokens(b.likely_report_theme)

    shared = len(dim_a & dim_b)
    union = max(1, len(dim_a | dim_b))
    overlap = shared / union
    combined_dim_count = len(dim_a | dim_b)
    theme_overlap = len(theme_a & theme_b) / max(1, len(theme_a | theme_b))

    score = 0.0
    score += combined_dim_count * 0.8
    score += min(shared, 5) * 1.4
    score += theme_overlap * 4.0
    if a.source_type != b.source_type:
        score += 2.0
    if 0.15 <= overlap <= 0.75:
        score += 3.0
    elif overlap > 0.90 and a.source_type == b.source_type:
        score -= 5.0
    return score


def choose_partner(seed: SourceRow, candidates: list[SourceRow]) -> SourceRow:
    return max(candidates, key=lambda candidate: complementarity_score(seed, candidate))


def cluster_quality(sources: list[SourceRow]) -> tuple[str, str]:
    canonical_ids = [source.canonical_source_id for source in sources]
    if len(set(canonical_ids)) != len(canonical_ids):
        return "needs_review", "duplicate canonical source detected inside cluster"

    dim_sets = [tokens(source.comparison_dimensions) for source in sources]
    union_dims = set().union(*dim_sets) if dim_sets else set()
    common_dims = set.intersection(*dim_sets) if dim_sets else set()
    source_types = {source.source_type for source in sources}

    if len(sources) < 2 or len(sources) > 3:
        return "needs_review", "cluster size outside 2-3 source target"
    if len(union_dims) < 6:
        return "needs_review", "comparison dimensions are too narrow"
    if len(common_dims) == 0:
        return "needs_review", "sources share weak comparison vocabulary"
    if len(sources) == 3 and len(source_types) == 1 and len(union_dims) < 10:
        return "needs_review", "sources may be semantically redundant"
    return "ready", "coherent COMPLETE sources with complementary comparison dimensions"


def concise_dimensions(sources: Iterable[SourceRow]) -> str:
    pieces: list[str] = []
    for source in sources:
        for part in source.comparison_dimensions.split(";"):
            part = part.strip()
            if part and part not in pieces:
                pieces.append(part)
    return "; ".join(pieces[:6])


def expected_output_type(domain: str) -> str:
    if domain == "Corporate performance and finance":
        return "comparative financial analysis"
    if domain == "AI and digital governance":
        return "comparative governance synthesis"
    if domain == "Economic outlook and development":
        return "comparative economic briefing"
    if domain == "Labor, education, and skills":
        return "comparative policy analysis"
    if domain == "Energy, climate, and environment":
        return "comparative trend analysis"
    if domain == "Health and social outcomes":
        return "evidence synthesis brief"
    return "comparative evidence synthesis"


def synthesis_intent(domain: str, dimensions: str, sources: list[SourceRow]) -> str:
    types = ", ".join(sorted({source.source_type for source in sources}))
    return (
        f"Synthesize {domain.lower()} evidence across {dimensions}, using {types} "
        "to identify contrasts, shared patterns, and report-ready implications."
    )


def cluster_rationale(sources: list[SourceRow], status_reason: str) -> str:
    source_types = ", ".join(sorted({source.source_type for source in sources}))
    ids = ", ".join(source.source_id for source in sources)
    return (
        f"{ids} are canonical COMPLETE sources in the same domain. The cluster combines "
        f"{source_types} with overlapping but not identical comparison dimensions; {status_reason}."
    )


def make_clusters_for_domain(domain: str, sources: list[SourceRow], start_idx: int) -> list[Cluster]:
    remaining = sorted(
        sources,
        key=lambda source: (
            source.source_type,
            source.comparison_dimensions,
            source.source_id,
        ),
    )
    clusters: list[Cluster] = []
    n = len(remaining)
    if n == 1:
        target_sizes = [1]
    elif n % 3 == 0:
        target_sizes = [3] * (n // 3)
    elif n % 3 == 1:
        # Avoid a singleton by ending with two pairs.
        target_sizes = [3] * max(0, (n // 3) - 1) + [2, 2]
    else:
        target_sizes = [3] * (n // 3) + [2]

    for target_size in target_sizes:
        seed = remaining.pop(0)
        group = [seed]
        while remaining and len(group) < target_size:
            partner = max(
                remaining,
                key=lambda candidate: sum(
                    complementarity_score(candidate, source) for source in group
                ),
            )
            remaining.remove(partner)
            group.append(partner)

        status, reason = cluster_quality(group)
        dimensions = concise_dimensions(group)
        clusters.append(
            Cluster(
                cluster_id="",
                sources=group,
                cluster_status=status,
                cluster_rationale=cluster_rationale(group, reason),
                expected_output_type=expected_output_type(domain),
                comparison_dimensions=dimensions,
                synthesis_intent=synthesis_intent(domain, dimensions, group),
            )
        )

    for offset, cluster in enumerate(clusters, start=start_idx):
        cluster.cluster_id = f"C{offset:03d}"
    return clusters


def build_clusters(sources: list[SourceRow]) -> list[Cluster]:
    by_domain: dict[str, list[SourceRow]] = defaultdict(list)
    for source in sources:
        by_domain[source.domain].append(source)

    clusters: list[Cluster] = []
    next_idx = 1
    for domain in sorted(by_domain, key=lambda key: (-len(by_domain[key]), key)):
        domain_clusters = make_clusters_for_domain(domain, by_domain[domain], next_idx)
        clusters.extend(domain_clusters)
        next_idx += len(domain_clusters)
    return clusters


def write_workbook(
    clusters: list[Cluster],
    output_path: Path,
    source_count: int,
    excluded_status_counts: Counter[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "clusters"

    headers = [
        "cluster_id",
        "canonical_source_ids",
        "source_ids",
        "domain",
        "cluster_rationale",
        "expected_output_type",
        "comparison_dimensions",
        "synthesis_intent",
        "cluster_status",
    ]
    ws.append(headers)

    for cluster in clusters:
        ws.append(
            [
                cluster.cluster_id,
                "; ".join(source.canonical_source_id for source in cluster.sources),
                "; ".join(source.source_id for source in cluster.sources),
                cluster.sources[0].domain if cluster.sources else "",
                cluster.cluster_rationale,
                cluster.expected_output_type,
                cluster.comparison_dimensions,
                cluster.synthesis_intent,
                cluster.cluster_status,
            ]
        )

    ready_fill = PatternFill("solid", fgColor="D9EAD3")
    review_fill = PatternFill("solid", fgColor="FCE5CD")
    for row in ws.iter_rows(min_row=2):
        status = row[8].value
        fill = ready_fill if status == "ready" else review_fill
        for cell in row:
            cell.value = clean_excel_value(cell.value)
            if isinstance(cell.value, str):
                cell.data_type = "s"
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 12,
        "B": 28,
        "C": 24,
        "D": 30,
        "E": 60,
        "F": 26,
        "G": 60,
        "H": 70,
        "I": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("summary")
    domain_counts = Counter(cluster.sources[0].domain for cluster in clusters if cluster.sources)
    status_counts = Counter(cluster.cluster_status for cluster in clusters)
    weak_clusters = [
        cluster.cluster_id for cluster in clusters if cluster.cluster_status == "needs_review"
    ]
    non_complete_excluded = sum(excluded_status_counts.values())

    summary_rows = [
        ("metric", "value"),
        ("number_of_clusters", len(clusters)),
        ("number_of_sources_used", source_count),
        ("number_of_non_complete_sources_excluded", non_complete_excluded),
        ("number_of_rejected_sources_excluded", excluded_status_counts.get("REJECTED", 0)),
        (
            "excluded_status_distribution",
            "; ".join(f"{k}: {v}" for k, v in excluded_status_counts.items()),
        ),
        ("weak_clusters_needing_manual_review", len(weak_clusters)),
        ("weak_cluster_ids", "; ".join(weak_clusters)),
        ("cluster_status_distribution", "; ".join(f"{k}: {v}" for k, v in status_counts.items())),
        ("domain_distribution", "; ".join(f"{k}: {v}" for k, v in domain_counts.items())),
    ]
    for row in summary_rows:
        summary.append(row)

    summary.append(())
    summary.append(("domain", "cluster_count"))
    for domain, count in domain_counts.most_common():
        summary.append((domain, count))

    for row in summary.iter_rows():
        for cell in row:
            cell.value = clean_excel_value(cell.value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in summary[1]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 100

    wb.save(output_path)


def run(input_path: Path, output_path: Path) -> int:
    sources, excluded_status_counts = load_complete_sources(input_path)
    clusters = build_clusters(sources)
    write_workbook(clusters, output_path, len(sources), excluded_status_counts)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    base_dir = Path(__file__).resolve().parent
    parser.add_argument("--input", type=Path, default=base_dir / "core_extraction.xlsx")
    parser.add_argument("--output", type=Path, default=base_dir / "core_clusters.xlsx")
    args = parser.parse_args()
    raise SystemExit(run(args.input, args.output))


if __name__ == "__main__":
    main()
