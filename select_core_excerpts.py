#!/usr/bin/env python3
"""Select source-grounded CORE benchmark excerpts from raw_text/*.txt."""

from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


WORD_RE = re.compile(r"\b[\w][\w'’.-]*\b", re.UNICODE)
SOURCE_ID_RE = re.compile(r"^(S\d{3})\.txt$", re.IGNORECASE)

SECTION_KEYWORDS = (
    "executive summary",
    "summary",
    "key findings",
    "main findings",
    "findings",
    "highlights",
    "overview",
    "conclusion",
    "conclusions",
    "recommendations",
    "results",
    "in brief",
    "key messages",
    "policy implications",
    "overview of the act",
    "policy background",
    "explanatory notes",
)

CONTENT_KEYWORDS = (
    "increase",
    "decrease",
    "growth",
    "decline",
    "higher",
    "lower",
    "compared",
    "comparison",
    "whereas",
    "while",
    "evidence",
    "data",
    "survey",
    "estimate",
    "estimated",
    "result",
    "results",
    "finding",
    "findings",
    "conclude",
    "conclusion",
    "impact",
    "risk",
    "trend",
    "average",
    "median",
    "share",
    "rate",
    "percent",
    "percentage",
    "billion",
    "million",
    "significant",
    "notably",
    "therefore",
    "however",
    "because",
)

BOILERPLATE_PATTERNS = (
    "all rights reserved",
    "creative commons",
    "isbn",
    "please cite",
    "photo credits",
    "copyright",
    "cookies",
    "skip to",
    "search legislation",
    "you are here",
    "table of contents",
    "contents overview",
    "skip to main content",
    "cookies on",
    "previous next",
    "more resources",
    "access essential accompanying documents",
    "forward-looking statements",
    "risks and uncertainties",
    "actual results",
    "differ materially",
    "cautionary statements",
    "unresolved staff comments",
    "risk factors",
    "safe harbor",
    "part i item",
    "item 1. business",
    "rights and permissions",
    "license",
    "electronically signed",
    "terms of use",
    "privacy",
    "accessibility",
    "contact us",
    "print this page",
    "download",
    "share this page",
    "official journal",
    "this document is made available",
    "without prejudice to the status",
)

LEGAL_BOILERPLATE_PATTERNS = (
    "be it enacted",
    "king's most excellent majesty",
    "lords spiritual and temporal",
    "short title",
    "commencement and extent",
    "this act extends",
    "this section applies",
    "subsection",
    "schedule",
)


@dataclass
class Candidate:
    start_line: int
    end_line: int
    text: str
    words: int
    score: float


@dataclass
class Selection:
    source_id: str
    excerpt_1: str = ""
    excerpt_2: str = ""
    excerpt_3: str = ""
    word_count: int = 0
    short_excerpt: bool = False
    notes: str = ""


def setup_logging(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("select_core_excerpts")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)
    return logger


def word_count(text: str) -> int:
    return len(WORD_RE.findall(text))


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def sanitize_for_excel(text: str) -> str:
    """Remove control characters that XLSX cells cannot store."""
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def is_table_like(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    words = word_count(stripped)
    tab_count = stripped.count("\t")
    numeric_tokens = len(re.findall(r"(?:\d+[.,]?)+%?|\$|€|£", stripped))
    if numeric_tokens >= 20:
        return True
    if re.search(r"\.{8,}", stripped):
        return True
    if tab_count >= 4 and words < 35:
        return True
    if tab_count >= 8:
        return True
    if numeric_tokens >= max(5, words // 2) and words < 40:
        return True
    return False


def line_penalty(line: str) -> float:
    lower = line.lower()
    penalty = 0.0
    if any(pattern in lower for pattern in BOILERPLATE_PATTERNS):
        penalty += 10.0
    if any(pattern in lower for pattern in LEGAL_BOILERPLATE_PATTERNS):
        penalty += 3.0
    if is_table_like(line):
        penalty += 4.0
    if re.fullmatch(r"[\W\d_]{1,20}", line.strip() or " "):
        penalty += 2.0
    if line.strip().startswith("====="):
        penalty += 8.0
    if re.search(r"\.{8,}", line):
        penalty += 8.0
    return penalty


def noisy_candidate_ratio(text: str) -> float:
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return 1.0
    noisy = 0
    for line in lines:
        lower = line.lower()
        if (
            any(pattern in lower for pattern in BOILERPLATE_PATTERNS)
            or is_table_like(line)
            or line.strip().startswith("=====")
            or re.search(r"\.{8,}", line)
        ):
            noisy += 1
    return noisy / len(lines)


def line_score(line: str) -> float:
    stripped = line.strip()
    if not stripped:
        return 0.0

    lower = stripped.lower()
    words = word_count(stripped)
    score = 0.0

    if any(keyword in lower for keyword in SECTION_KEYWORDS):
        score += 12.0
    score += sum(1.0 for keyword in CONTENT_KEYWORDS if keyword in lower)
    score += min(4.0, len(re.findall(r"\b\d+(?:[.,]\d+)?%?\b", stripped)) * 0.7)

    if words >= 18 and re.search(r"[.!?]$", stripped):
        score += 1.5
    if words < 4:
        score -= 1.5

    score -= line_penalty(stripped)
    return score


def find_section_boosts(lines: list[str]) -> list[float]:
    boosts = [0.0] * len(lines)
    section_positions: list[int] = []
    for idx, line in enumerate(lines):
        lower = line.strip().lower()
        if any(keyword == lower or keyword in lower for keyword in SECTION_KEYWORDS):
            section_positions.append(idx)

    for pos in section_positions:
        for idx in range(pos, min(len(lines), pos + 90)):
            distance = idx - pos
            boosts[idx] += max(0.0, 8.0 - distance * 0.12)
    return boosts


def build_candidates(text: str) -> list[Candidate]:
    lines = text.splitlines()
    boosts = find_section_boosts(lines)
    candidates: list[Candidate] = []

    start = 0
    while start < len(lines):
        while start < len(lines) and not lines[start].strip():
            start += 1
        if start >= len(lines):
            break

        collected: list[str] = []
        total_words = 0
        total_score = 0.0
        end = start

        while end < len(lines) and total_words < 450:
            line = lines[end]
            stripped = line.strip()
            if stripped.startswith("====="):
                end += 1
                continue
            if stripped:
                collected.append(line.rstrip())
                total_words += word_count(stripped)
                total_score += line_score(stripped) + boosts[end]
            elif total_words >= 180:
                break
            elif collected:
                collected.append("")

            if total_words >= 220 and re.search(r"[.!?]\s*$", stripped):
                break
            end += 1

        candidate_text = "\n".join(collected).strip()
        candidate_words = word_count(candidate_text)
        noise = noisy_candidate_ratio(candidate_text)
        if candidate_words >= 40 and noise < 0.35:
            density = total_score / max(candidate_words, 1)
            # Prefer substantive middle/front-matter sections without overselecting covers.
            position_bonus = 0.7 if start > len(lines) * 0.02 else -0.5
            if noise > 0.15:
                position_bonus -= 2.0
            candidates.append(
                Candidate(
                    start_line=start,
                    end_line=end,
                    text=candidate_text,
                    words=candidate_words,
                    score=density + position_bonus,
                )
            )

        start = max(end + 1, start + 10)

    return candidates


def trim_to_word_limit(text: str, max_words: int) -> str:
    if word_count(text) <= max_words:
        return text

    lines: list[str] = []
    total = 0
    for line in text.splitlines():
        line_words = word_count(line)
        if total + line_words > max_words:
            break
        lines.append(line)
        total += line_words
    if lines:
        return "\n".join(lines).strip()

    words = WORD_RE.findall(text)
    limit_word = words[max_words - 1]
    idx = text.find(limit_word)
    return text[: idx + len(limit_word)].strip()


def overlaps(a: Candidate, b: Candidate) -> bool:
    return not (a.end_line < b.start_line or b.end_line < a.start_line)


def select_excerpts(source_id: str, text: str, logger: logging.Logger) -> Selection:
    candidates = sorted(build_candidates(text), key=lambda c: c.score, reverse=True)
    selected: list[Candidate] = []
    total_words = 0

    for candidate in candidates:
        if any(overlaps(candidate, existing) for existing in selected):
            continue
        selected.append(candidate)
        total_words += candidate.words
        if len(selected) >= 3:
            break
        if total_words >= 1300:
            break

    if not selected:
        fallback = trim_to_word_limit(text.strip(), 500)
        logger.warning("no scored candidates for %s; using fallback", source_id)
        return Selection(
            source_id=source_id,
            excerpt_1=fallback,
            word_count=word_count(fallback),
            short_excerpt=word_count(fallback) < 500,
            notes="fallback_no_candidates",
        )

    selected = sorted(selected[:3], key=lambda c: c.start_line)
    excerpts = [candidate.text for candidate in selected]
    total_words = sum(word_count(excerpt) for excerpt in excerpts)

    if total_words > 1500:
        remaining = 1500
        trimmed: list[str] = []
        for excerpt in excerpts:
            trimmed_excerpt = trim_to_word_limit(excerpt, remaining)
            if trimmed_excerpt:
                trimmed.append(trimmed_excerpt)
                remaining -= word_count(trimmed_excerpt)
            else:
                trimmed.append("")
        excerpts = trimmed
        total_words = sum(word_count(excerpt) for excerpt in excerpts)

    short = total_words < 500
    if short:
        logger.warning("%s has short excerpt: %d words", source_id, total_words)

    while len(excerpts) < 3:
        excerpts.append("")

    return Selection(
        source_id=source_id,
        excerpt_1=excerpts[0],
        excerpt_2=excerpts[1],
        excerpt_3=excerpts[2],
        word_count=total_words,
        short_excerpt=short,
        notes="" if not short else "short_excerpt",
    )


def write_workbook(selections: list[Selection], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "core_extraction"

    headers = [
        "source_id",
        "excerpt_1",
        "excerpt_2",
        "excerpt_3",
        "word_count",
        "short_excerpt",
        "notes",
    ]
    ws.append(headers)

    for selection in selections:
        ws.append(
            [
                selection.source_id,
                sanitize_for_excel(selection.excerpt_1),
                sanitize_for_excel(selection.excerpt_2),
                sanitize_for_excel(selection.excerpt_3),
                selection.word_count,
                selection.short_excerpt,
                selection.notes,
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = {
        "A": 12,
        "B": 80,
        "C": 80,
        "D": 80,
        "E": 12,
        "F": 14,
        "G": 24,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(output_path)


def run(raw_text_dir: Path, output_path: Path, log_path: Path) -> int:
    logger = setup_logging(log_path)
    raw_text_dir = raw_text_dir.resolve()
    output_path = output_path.resolve()

    if not raw_text_dir.is_dir():
        logger.error("raw_text directory does not exist: %s", raw_text_dir)
        return 1

    paths = sorted(
        path for path in raw_text_dir.glob("*.txt") if SOURCE_ID_RE.match(path.name)
    )
    if not paths:
        logger.error("no raw_text/*.txt files found in %s", raw_text_dir)
        return 1

    selections: list[Selection] = []
    for path in paths:
        source_id = path.stem.upper()
        try:
            text = path.read_text(encoding="utf-8")
            selections.append(select_excerpts(source_id, text, logger))
            logger.info("selected excerpts for %s", source_id)
        except Exception as exc:
            logger.exception("failed to process %s: %s", path, exc)
            selections.append(
                Selection(
                    source_id=source_id,
                    short_excerpt=True,
                    notes=f"failed: {exc}",
                )
            )

    write_workbook(selections, output_path)
    short_count = sum(1 for selection in selections if selection.short_excerpt)
    logger.info(
        "wrote %s (%d sources, %d short_excerpt)",
        output_path,
        len(selections),
        short_count,
    )
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--raw-text-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "raw_text",
        help="Directory containing raw_text/S###.txt files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "core_extraction.xlsx",
        help="Output Excel workbook path.",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=Path(__file__).resolve().parent / "core_extraction.log",
        help="Log path.",
    )
    args = parser.parse_args()
    raise SystemExit(run(args.raw_text_dir, args.output, args.log))


if __name__ == "__main__":
    main()
